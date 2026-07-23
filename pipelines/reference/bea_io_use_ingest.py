"""BEA Input-Output Use Tables (Supply-Use Framework) → Gen-3 Lance SoR.

Directive: hq/directives/2026-07-11-bea-io-use-contingent-labor-ingest.md. Closes the third
leg of the labor-sourcing decomposition (self-perform / subcontract / **contingent**). A
prime's staffing-agency spend is a vendor purchase, not a reported subaward, so FPDS/FSRS are
structurally blind to it. The BEA use table is the only public, economy-wide measurement of
that channel: for every industry, the dollars of **Employment services** (NAICS 5613 —
temp/staffing/PEO) consumed as a purchased input. Landing the full matrix yields every other
purchased-input intensity for free, plus the workbook's own BEA↔NAICS concordance.

Pattern A direct hydration from ONE keyless static zip (AllTablesSUP.zip, 20,439,248 B,
browser-UA; the BEA API is key-gated and no key exists in Doppler — do not use it). Two of the
15 members are in scope; the CxC/IxC/IxI total-requirements and Supply tables are not.

  --stream detail       → active/bea_io_use_detail/            ~159.5K rows
      Use_SUT_Framework_2017_DET.xlsx, sheets 2007/2012/2017 (detail grain exists ONLY for the
      3 benchmark years — the annual series is Summary grain; that asymmetry is structural).
      436 x 427 per sheet, $Millions. Row 5 = industry names, row 6 = industry codes, data
      row 7+. Col A = commodity code, col B = commodity description, cols C+ = using industries
      then T001 / 19 final-demand columns / T019.
  --stream summary      → active/bea_io_use_summary_annual/    ~206.2K rows
      Use_Tables_Supply-Use_Framework_1997-2023_Summary.xlsx, one sheet per year 1997-2023 (27),
      90 x 94 at 71-industry grain. **Header order is INVERTED vs the detail workbook**: row 6 =
      industry codes, row 7 = industry names, data row 8+. Verified 2026-07-23; §2.4 of the
      directive flagged this workbook's internals as the one un-pre-verified structure.
  --stream concordance  → active/bea_sut_naics_concordance/    643 rows
      The detail workbook's own `NAICS Codes` sheet: a hierarchical outline (one row per BEA
      code at whichever of the 4 levels it occupies, title in the column immediately right of
      the code, `Related 2017 NAICS Codes` present on detail rows only).
  --stream derived      → active/bea_contingent_labor_intake/  3,067 rows
      industry x year contingent-labor intake: employment-services input over total industry
      output and over total intermediate inputs.

DATASET-NAME DEVIATION FROM THE DIRECTIVE (deliberate, load-bearing)
    §4 names the concordance stream `bea_naics_concordance`. That name is ALREADY LIVE from the
    predecessor labor-share stack (pipelines/reference/materialize_bea_naics_concordance.py,
    499 rows) built from a DIFFERENT source — the standalone BEA-Industry-and-Commodity-Codes-
    and-NAICS-Concordance.xlsx, whose layout is one row per NAICS code with all five BEA levels
    (incl. GO Detail) filled across the row. Writing this stream to that URI with mode=overwrite
    would clobber a richer, differently-grained dataset and violate the directive's own §2
    constraint that the two modules' datasets stay disjoint. This stream therefore lands at
    `bea_sut_naics_concordance` — the Supply-Use-workbook vintage, hierarchical-outline grain.
    The two are complements, not substitutes: use bea_naics_concordance to resolve a NAICS code
    to its BEA levels, and bea_sut_naics_concordance to resolve any code appearing in THIS
    module's use matrices (its bea_code values are exactly the row/column codes landed here).

FIDELITY (raw stays lossless)
    Commodity/industry codes and names land verbatim (numeric-looking codes are rendered
    without a float artifact: 111200 not 111200.0). Values are $Millions, DOUBLE. Published
    zeros are real observations and land as 0.0 — never conflated with absence. Blank cells
    (the detail workbook's "no flow" encoding, 71% of that matrix) are not landed. The summary
    workbook has ZERO blanks: every cell is numeric or the literal marker `...`, which BEA's own
    legend on the `NAICS Codes` sheet defines as **"Not applicable"** — the summary workbook's
    encoding of the same "no flow" the detail workbook writes as blank. It is NOT a disclosure
    suppression (BEA does not suppress in the national IO accounts at 71-industry grain). Those
    cells land with value_musd NULL and the marker preserved verbatim in `value_marker`;
    `suppressed` is carried per the directive's §4 column contract and is exactly
    `value_marker IS NOT NULL`. Read it as "cell carried a non-numeric published marker", not
    as "BEA withheld this number". Filter `value_musd IS NOT NULL` for flows in either stream.
    Row/column kinds are ADDITIONAL classifier columns, never a filter applied at write time:
    the T*/V* rows (total intermediate inputs, compensation of employees, gross operating
    surplus, total industry output) are the denominators every intensity ratio needs, and the
    F* columns are final demand.

CONTROL PLANE
    Ledger: ops.labor_share_runs (shared with the predecessor stack, append-only; one row per
    Lance dataset). Every stream runs a fail-closed structural gate BEFORE writing: the
    industry/commodity/final-demand column and row censuses must match the verified shape, or
    the parse raises rather than landing a silently-reshaped matrix.

    doppler run -p core-x -c prd -- uv run --with pylance --with pyarrow --with openpyxl \\
      --with requests --with boto3 --with 'psycopg[binary]' \\
      python -m pipelines.reference.bea_io_use_ingest --stream all --smoke   # throwaway URIs
    ... python -m pipelines.reference.bea_io_use_ingest --stream all                    # full
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import os
import os.path
import re
import tempfile
import zipfile

# Fleet R2/index/ledger plumbing — reused verbatim, never reimplemented.
from pipelines.bls.ingest import _build_indexes, _storage_options  # noqa: E402
from pipelines.reference.labor_share_ingest import (  # noqa: E402
    BUCKET,
    _delete_prefix,
    _download,
    _record_run,
    _s,
    _write_lance,
)

SOURCE_TAG = "bea_io_use_ingest"
SOURCE_URL = "https://apps.bea.gov/industry/iTables%20Static%20Files/AllTablesSUP.zip"
ZIP_BYTES_EXPECTED = 20_439_248

DETAIL_MEMBER = "Use_SUT_Framework_2017_DET.xlsx"
SUMMARY_MEMBER = "Use_Tables_Supply-Use_Framework_1997-2023_Summary.xlsx"
CONCORDANCE_SHEET = "NAICS Codes"

DETAIL_URI = os.environ.get("BEA_IO_USE_DETAIL_URI", f"s3://{BUCKET}/active/bea_io_use_detail/")
SUMMARY_URI = os.environ.get("BEA_IO_USE_SUMMARY_URI",
                             f"s3://{BUCKET}/active/bea_io_use_summary_annual/")
CONCORDANCE_URI = os.environ.get("BEA_SUT_CONCORDANCE_URI",
                                 f"s3://{BUCKET}/active/bea_sut_naics_concordance/")
DERIVED_URI = os.environ.get("BEA_CONTINGENT_INTAKE_URI",
                             f"s3://{BUCKET}/active/bea_contingent_labor_intake/")

# ── verified workbook geometry (probed live 2026-07-23; asserted fail-closed at parse) ──
# name_row/code_row/data_row are 1-based worksheet rows; *_col are 0-based tuple indexes.
DETAIL_SPEC = {
    "member": DETAIL_MEMBER, "grain": "detail",
    "name_row": 5, "code_row": 6, "data_row": 7, "code_col": 0, "desc_col": 1, "first_val_col": 2,
    "sheets": ("2007", "2012", "2017"),
    "census": {"industry": 402, "final_demand": 19, "total": 2, "commodity": 402, "total_or_va": 9},
    # Blank == "no flow", so the detail matrix is sparse (29.1% dense) and only a floor applies.
    "cells_per_year_min": 50_000, "cells_per_year_exact": None,
}
SUMMARY_SPEC = {
    "member": SUMMARY_MEMBER, "grain": "summary",
    # header order is INVERTED vs detail: codes on row 6, names on row 7.
    "name_row": 7, "code_row": 6, "data_row": 8, "code_col": 0, "desc_col": 1, "first_val_col": 2,
    "sheets": tuple(str(y) for y in range(1997, 2024)),
    "census": {"industry": 71, "final_demand": 19, "total": 2, "commodity": 73, "total_or_va": 10},
    # The summary matrix has ZERO blanks — every cell is numeric or the `...` not-applicable
    # marker — so it lands as a full rectangle: (73 commodity + 10 total/VA) rows x (71 industry
    # + 19 final-demand + 2 total) columns = 7,636 cells per year, every year. Exact equality is
    # a far stronger completeness assertion than any floor.
    "cells_per_year_min": None, "cells_per_year_exact": 83 * 92,
}

# Total / value-added code shapes. Rows: T005 total intermediate inputs, V001/V00100 compensation
# of employees, T00OTOP other taxes on production, T00OSUB less other subsidies, V003/V00300 gross
# operating surplus, VABAS/VAPRO value added, T018 total industry output, T00TOP/T00SUB taxes and
# subsidies on products. Columns: T001 total intermediate, T019 total use of products.
_TOTAL_CODE_RX = re.compile(r"^(T\d|T00[A-Z]|V\d|VABAS|VAPRO)", re.I)
_FINAL_DEMAND_RX = re.compile(r"^F\d", re.I)

# The contingent-labor target row: detail carries the 5613 commodity outright; at summary grain
# BEA dissolves it into the whole administrative-and-support aggregate (561), so the annual
# series is a PROXY for the staffing channel, not a measurement of it.
EMP_SVCS_DETAIL_CODE = "561300"
EMP_SVCS_SUMMARY_CODE = "561"
TOTAL_INTERMEDIATE_ROW = "T005"
TOTAL_OUTPUT_ROW = "T018"

# ── Validation Gate (directive §8). Deviations from the directive's stated bounds are
# annotated inline; each is a directive ESTIMATE corrected against the live workbook, never a
# loosened correctness check. The real completeness proof is GATE_RECON_MIN (§8.5).
GATE_DETAIL_YEARS = {"2007", "2012", "2017"}
GATE_MIN_PURCHASING_INDUSTRIES = 50
GATE_2017_EMP_SVCS_RANGE = (150_000.0, 500_000.0)
# §8.1's "per-year non-null cells >= 60,000" is a DETAIL-stream bound (§8.2 sets no cell floor
# for summary). The live detail matrix is 29.1% dense — 53,115 / 53,201 / 53,222 across
# 2007/2012/2017 — so 60,000 is unreachable and was an estimate of matrix density, not a
# measured floor. Each spec carries its own bound: a floor for the sparse detail matrix, exact
# rectangularity for the dense summary one (see the specs above).
GATE_SUMMARY_YEARS = {str(y) for y in range(1997, 2024)}
GATE_CODE_DRIFT_TOLERANCE = 2
GATE_MIN_CONCORDANCE_ROWS = 380
# §8.4 bounds intake_share_of_output at (0, 0.15]. That holds at detail grain (max 0.0967).
# The summary series measures 561 — the ENTIRE administrative-and-support aggregate, of which
# employment services is one part — so it structurally runs higher (max 0.1745, pipeline
# transportation 1998). Detail keeps the directive's bound; the proxy gets its own.
GATE_DETAIL_SHARE_MAX = 0.15
GATE_SUMMARY_SHARE_MAX = 0.20
GATE_RECON_MIN = 0.98
GATE_STAFFING_HEAVY_RX = re.compile(
    r"employment services|business support|facilities support|administrative|"
    r"hospital|health care|nursing|transportation|warehous|courier|transit",
    re.I,
)


# ── source zip (downloaded once per process; --stream all reuses it) ────────────────
_ZIP_CACHE: dict[str, zipfile.ZipFile] = {}


def _zip() -> zipfile.ZipFile:
    if "zf" in _ZIP_CACHE:
        return _ZIP_CACHE["zf"]
    cached = os.path.join(tempfile.gettempdir(), "bea_AllTablesSUP.zip")
    if os.path.isfile(cached) and os.path.getsize(cached) == ZIP_BYTES_EXPECTED:
        with open(cached, "rb") as fh:
            raw = fh.read()
        print(f"  zip: reusing scratch copy {cached} ({len(raw):,} B)", flush=True)
    else:
        raw = _download(SOURCE_URL, timeout=600)
        print(f"  zip: downloaded {len(raw):,} B from {SOURCE_URL}", flush=True)
        if len(raw) != ZIP_BYTES_EXPECTED:
            # Upstream is a static file re-verified 2026-07-23; a size change means the vintage
            # moved and the pre-verified geometry can no longer be assumed.
            raise RuntimeError(
                f"AllTablesSUP.zip is {len(raw):,} B, expected {ZIP_BYTES_EXPECTED:,} B — "
                "upstream vintage changed; re-probe the workbook geometry before ingesting.")
        try:
            with open(cached, "wb") as fh:
                fh.write(raw)
        except OSError:
            pass
    zf = zipfile.ZipFile(io.BytesIO(raw))
    names = set(zf.namelist())
    missing = {DETAIL_MEMBER, SUMMARY_MEMBER} - names
    if missing:
        raise RuntimeError(f"AllTablesSUP.zip is missing target member(s): {sorted(missing)}")
    _ZIP_CACHE["zf"] = zf
    return zf


def _workbook(member: str):
    import openpyxl

    return openpyxl.load_workbook(io.BytesIO(_zip().read(member)), read_only=True, data_only=True)


# ── cell coercion ──────────────────────────────────────────────────────────────────
def _code(v) -> str | None:
    """Code cell → verbatim string; ints/whole floats render without a decimal artifact."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip() or None


def _cell(v) -> tuple[float | None, str | None]:
    """Matrix cell → (value_musd, value_marker). Blank → (None, None) and is not landed.
    A published zero is a real observation and returns (0.0, None)."""
    if v is None or isinstance(v, bool):
        return (None, None)
    if isinstance(v, (int, float)):
        return (float(v), None)
    s = str(v).strip()
    if not s:
        return (None, None)
    try:
        return (float(s.replace(",", "")), None)
    except ValueError:
        return (None, s)


def _col_kind(code: str) -> str:
    if _FINAL_DEMAND_RX.match(code):
        return "final_demand"
    if _TOTAL_CODE_RX.match(code):
        return "total"
    return "industry"


def _row_kind(code: str) -> str:
    return "total_or_va" if _TOTAL_CODE_RX.match(code) else "commodity"


def _matrix_schema():
    import pyarrow as pa

    return pa.schema([
        ("year", pa.int32()),
        ("commodity_code", pa.string()), ("commodity_name", pa.string()),
        ("industry_code", pa.string()), ("industry_name", pa.string()),
        ("col_kind", pa.string()), ("row_kind", pa.string()),
        ("value_musd", pa.float64()), ("suppressed", pa.bool_()), ("value_marker", pa.string()),
        ("grain", pa.string()), ("source", pa.string()), ("source_file", pa.string()),
        ("source_url", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


# ── shared melt: one sheet → long-form rows, fail-closed on a reshaped matrix ───────
def _sheet_header(ws, spec: dict):
    """(columns, rows) where columns = [(idx, code, name, col_kind)] and rows are raw tuples
    from the first data row onward. Raises if the column census drifts from the verified shape."""
    grid = list(ws.iter_rows(min_row=1, values_only=True))
    codes_r = grid[spec["code_row"] - 1]
    names_r = grid[spec["name_row"] - 1]
    cols = []
    for j in range(spec["first_val_col"], len(codes_r)):
        c = _code(codes_r[j])
        if not c:
            continue  # trailing padding columns
        cols.append((j, c, _s(names_r[j]) if j < len(names_r) else None, _col_kind(c)))
    census = {k: sum(1 for _j, _c, _n, k2 in cols if k2 == k)
              for k in ("industry", "final_demand", "total")}
    want = {k: spec["census"][k] for k in census}
    if census != want:
        raise RuntimeError(
            f"{ws.title}: column census {census} != verified {want} — the workbook was reshaped; "
            "re-probe geometry before ingesting.")
    return cols, grid[spec["data_row"] - 1:]


def _melt_sheet(ws, spec: dict, year: int, ingested_at) -> list[dict]:
    cols, body = _sheet_header(ws, spec)
    ccol, dcol = spec["code_col"], spec["desc_col"]
    grain, member = spec["grain"], spec["member"]
    src = f"{SOURCE_TAG}:{grain}"
    out: list[dict] = []
    row_census = {"commodity": 0, "total_or_va": 0}
    for r in body:
        rcode = _code(r[ccol]) if ccol < len(r) else None
        rname = _s(r[dcol]) if dcol < len(r) else None
        # Real rows carry BOTH a code and a description. This is what drops BEA's trailing
        # "Note.  Detail may not add to total due to rounding." line, which otherwise parses
        # as a commodity code.
        if not rcode or not rname:
            continue
        rkind = _row_kind(rcode)
        row_census[rkind] += 1
        for j, icode, iname, ckind in cols:
            val, marker = _cell(r[j]) if j < len(r) else (None, None)
            if val is None and marker is None:
                continue
            out.append({
                "year": year, "commodity_code": rcode, "commodity_name": rname,
                "industry_code": icode, "industry_name": iname,
                "col_kind": ckind, "row_kind": rkind,
                "value_musd": val, "suppressed": marker is not None, "value_marker": marker,
                "grain": grain, "source": src, "source_file": member,
                "source_url": SOURCE_URL, "ingested_at": ingested_at,
            })
    want = {k: spec["census"][k] for k in row_census}
    if row_census != want:
        raise RuntimeError(
            f"{ws.title}: row census {row_census} != verified {want} — the workbook was "
            "reshaped; re-probe geometry before ingesting.")
    return out


def _run_matrix_stream(spec: dict, dataset: str, uri: str, storage_options: dict) -> dict:
    import pyarrow as pa

    started_at = dt.datetime.now(dt.timezone.utc)
    ingested_at = started_at
    status, error_text, built = "error", None, []
    rows: list[dict] = []
    cov: dict = {}
    per_year: dict[str, int] = {}
    code_sets: dict[str, tuple[frozenset, frozenset]] = {}
    try:
        wb = _workbook(spec["member"])
        missing = set(spec["sheets"]) - set(wb.sheetnames)
        if missing:
            raise RuntimeError(f"{spec['member']}: missing sheet(s) {sorted(missing)}")
        for sheet in spec["sheets"]:
            ws = wb[sheet]
            got = _melt_sheet(ws, spec, int(sheet), ingested_at)
            per_year[sheet] = len(got)
            code_sets[sheet] = (frozenset(r["commodity_code"] for r in got),
                                frozenset(r["industry_code"] for r in got))
            rows.extend(got)
            print(f"  [{spec['grain']}] {sheet}: {len(got):,} cells", flush=True)
        wb.close()

        _gate_matrix(spec, rows, per_year, code_sets)

        _write_lance(pa.Table.from_pylist(rows, schema=_matrix_schema()), uri, storage_options)
        built = _build_indexes(uri, btree=["commodity_code", "industry_code"],
                               bitmap=["year", "row_kind", "col_kind", "grain"], so=storage_options)
        status = "success"
    except Exception as exc:  # noqa: BLE001
        error_text = str(exc)
        print(f"FATAL {dataset}: {exc}", flush=True)
        raise
    finally:
        completed_at = dt.datetime.now(dt.timezone.utc)
        emp_code = (EMP_SVCS_DETAIL_CODE if spec["grain"] == "detail" else EMP_SVCS_SUMMARY_CODE)
        cov = {
            "rows": len(rows), "rows_per_year": per_year,
            "years": sorted(per_year),
            "valued_cells": sum(1 for r in rows if r["value_musd"] is not None),
            "marker_cells": sum(1 for r in rows if r["value_marker"] is not None),
            "zero_cells": sum(1 for r in rows if r["value_musd"] == 0.0),
            "distinct_commodities": len({r["commodity_code"] for r in rows}),
            "distinct_industries": len({r["industry_code"] for r in rows
                                        if r["col_kind"] == "industry"}),
            "emp_svcs_code": emp_code,
            "emp_svcs_intermediate_musd_by_year": {
                y: round(sum(r["value_musd"] for r in rows
                             if r["year"] == int(y) and r["commodity_code"] == emp_code
                             and r["col_kind"] == "industry" and r["value_musd"] is not None), 1)
                for y in sorted(per_year)},
        }
        _record_run(dataset, uri, SOURCE_URL, len(rows), built, cov, status, error_text,
                    started_at, completed_at)
        print(f"{dataset.upper()} SUMMARY: rows={cov.get('rows')} "
              f"valued={cov.get('valued_cells')} markers={cov.get('marker_cells')} "
              f"status={status}", flush=True)
    return {"dataset": dataset, "rows": len(rows), "indexes": built,
            "coverage": cov, "status": status}


def _gate_matrix(spec: dict, rows: list[dict], per_year: dict[str, int],
                 code_sets: dict[str, tuple[frozenset, frozenset]]) -> None:
    """Directive §8.1 / §8.2 / §8.5 — fail closed before anything is written."""
    grain = spec["grain"]
    want_years = GATE_DETAIL_YEARS if grain == "detail" else GATE_SUMMARY_YEARS
    got_years = set(per_year)
    if got_years != want_years:
        raise RuntimeError(f"GATE {grain}: years {sorted(got_years)} != {sorted(want_years)}")

    emp_code = EMP_SVCS_DETAIL_CODE if grain == "detail" else EMP_SVCS_SUMMARY_CODE
    for y in sorted(per_year):
        buying = {r["industry_code"] for r in rows
                  if r["year"] == int(y) and r["commodity_code"] == emp_code
                  and r["col_kind"] == "industry" and r["value_musd"]}
        if len(buying) < GATE_MIN_PURCHASING_INDUSTRIES:
            raise RuntimeError(f"GATE {grain} {y}: only {len(buying)} industries purchase "
                               f"{emp_code} (need >= {GATE_MIN_PURCHASING_INDUSTRIES})")
        floor, exact = spec["cells_per_year_min"], spec["cells_per_year_exact"]
        if floor is not None and per_year[y] < floor:
            raise RuntimeError(f"GATE {grain} {y}: {per_year[y]:,} cells (need >= {floor:,})")
        if exact is not None and per_year[y] != exact:
            raise RuntimeError(f"GATE {grain} {y}: {per_year[y]:,} cells != the full rectangle "
                               f"{exact:,} — the matrix is not dense as verified")

    if grain == "detail":
        total_2017 = sum(r["value_musd"] for r in rows
                         if r["year"] == 2017 and r["commodity_code"] == EMP_SVCS_DETAIL_CODE
                         and r["col_kind"] == "industry" and r["value_musd"] is not None)
        lo, hi = GATE_2017_EMP_SVCS_RANGE
        if not lo <= total_2017 <= hi:
            raise RuntimeError(f"GATE detail: 2017 economy-wide employment-services intermediate "
                               f"purchases ${total_2017:,.0f}M outside [{lo:,.0f}, {hi:,.0f}]")
    else:
        base_c, base_i = code_sets["2017"]
        for y, (cs, isx) in code_sets.items():
            drift = len(cs ^ base_c) + len(isx ^ base_i)
            if drift > GATE_CODE_DRIFT_TOLERANCE:
                raise RuntimeError(f"GATE summary {y}: code-set drift {drift} vs 2017 "
                                   f"(tolerance {GATE_CODE_DRIFT_TOLERANCE})")

    # §8.5 reconciliation — per (year, industry), the landed commodity inputs must not exceed
    # the published total intermediate inputs by more than the rounding tolerance.
    landed: dict[tuple[int, str], float] = {}
    published: dict[tuple[int, str], float] = {}
    for r in rows:
        if r["col_kind"] != "industry" or r["value_musd"] is None:
            continue
        key = (r["year"], r["industry_code"])
        if r["row_kind"] == "commodity":
            landed[key] = landed.get(key, 0.0) + r["value_musd"]
        elif r["commodity_code"] == TOTAL_INTERMEDIATE_ROW:
            published[key] = r["value_musd"]
    ratios = [published[k] / landed[k] for k in published if landed.get(k)]
    if not ratios:
        raise RuntimeError(f"GATE {grain}: no reconcilable (year, industry) pairs")
    if min(ratios) < GATE_RECON_MIN:
        raise RuntimeError(f"GATE {grain}: reconciliation min {min(ratios):.6f} < "
                           f"{GATE_RECON_MIN} — the melt lost material rows")
    print(f"  GATE {grain} ✓ years={len(per_year)} recon=[{min(ratios):.6f}, {max(ratios):.6f}] "
          f"n={len(ratios)}", flush=True)


def run_detail(*, storage_options: dict, uri: str = DETAIL_URI) -> dict:
    return _run_matrix_stream(DETAIL_SPEC, "bea_io_use_detail", uri, storage_options)


def run_summary(*, storage_options: dict, uri: str = SUMMARY_URI) -> dict:
    return _run_matrix_stream(SUMMARY_SPEC, "bea_io_use_summary_annual", uri, storage_options)


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM 3 — the detail workbook's own BEA ↔ 2017 NAICS concordance
# ═══════════════════════════════════════════════════════════════════════════════════
# Hierarchical outline: the code sits in the column for its level and the title in the very
# next column. `Related 2017 NAICS Codes` is populated on detail rows only. Codes repeat across
# levels where a parent has a single child (e.g. summary 211 and u.summary 211), so bea_code is
# NOT unique — (bea_level, bea_code) is.
_CONC_HEADER_ROW = 5
_CONC_LEVELS = {0: "sector", 1: "summary", 2: "u_summary", 3: "detail"}
_CONC_NOTES_COL = 5
_CONC_NAICS_COL = 6


def _concordance_schema():
    import pyarrow as pa

    return pa.schema([
        ("bea_code", pa.string()), ("bea_title", pa.string()), ("bea_level", pa.string()),
        ("naics_ranges", pa.string()), ("notes", pa.string()),
        ("sheet_source", pa.string()), ("source", pa.string()), ("source_file", pa.string()),
        ("source_url", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


def run_concordance(*, storage_options: dict, uri: str = CONCORDANCE_URI) -> dict:
    import pyarrow as pa

    started_at = dt.datetime.now(dt.timezone.utc)
    ingested_at = started_at
    status, error_text, built = "error", None, []
    rows: list[dict] = []
    cov: dict = {}
    try:
        wb = _workbook(DETAIL_MEMBER)
        if CONCORDANCE_SHEET not in wb.sheetnames:
            raise RuntimeError(f"{DETAIL_MEMBER}: sheet {CONCORDANCE_SHEET!r} absent")
        ws = wb[CONCORDANCE_SHEET]
        for r in ws.iter_rows(min_row=_CONC_HEADER_ROW + 1, values_only=True):
            for j, level in _CONC_LEVELS.items():
                code = _code(r[j]) if j < len(r) else None
                if not code:
                    continue
                title = _s(r[j + 1]) if j + 1 < len(r) else None
                # BEA's trailing legend/footnote block sits in column A with no title beside it
                # ("n.a. Not applicable.", the asterisk note, the four dagger notes). A real
                # concordance row always carries a title.
                if title:
                    rows.append({
                        "bea_code": code, "bea_title": title, "bea_level": level,
                        "naics_ranges": _code(r[_CONC_NAICS_COL]) if len(r) > _CONC_NAICS_COL else None,
                        "notes": _s(r[_CONC_NOTES_COL]) if len(r) > _CONC_NOTES_COL else None,
                        "sheet_source": CONCORDANCE_SHEET,
                        "source": f"{SOURCE_TAG}:concordance", "source_file": DETAIL_MEMBER,
                        "source_url": SOURCE_URL, "ingested_at": ingested_at,
                    })
                break
        wb.close()

        if len(rows) < GATE_MIN_CONCORDANCE_ROWS:
            raise RuntimeError(f"GATE concordance: {len(rows)} rows < {GATE_MIN_CONCORDANCE_ROWS}")
        target = [r for r in rows if r["bea_code"] == EMP_SVCS_DETAIL_CODE]
        if not target:
            raise RuntimeError(f"GATE concordance: {EMP_SVCS_DETAIL_CODE} absent")
        print(f"  GATE concordance ✓ rows={len(rows)} 561300 -> "
              f"naics {target[0]['naics_ranges']!r}", flush=True)

        _write_lance(pa.Table.from_pylist(rows, schema=_concordance_schema()), uri, storage_options)
        built = _build_indexes(uri, btree=["bea_code"], bitmap=["bea_level"], so=storage_options)
        status = "success"
    except Exception as exc:  # noqa: BLE001
        error_text = str(exc)
        print(f"FATAL bea_sut_naics_concordance: {exc}", flush=True)
        raise
    finally:
        completed_at = dt.datetime.now(dt.timezone.utc)
        cov = {
            "rows": len(rows),
            "by_level": {lvl: sum(1 for r in rows if r["bea_level"] == lvl)
                         for lvl in _CONC_LEVELS.values()},
            "with_naics_ranges": sum(1 for r in rows if r["naics_ranges"]),
            "with_notes": sum(1 for r in rows if r["notes"]),
            "distinct_codes": len({r["bea_code"] for r in rows}),
            "emp_svcs_naics": next((r["naics_ranges"] for r in rows
                                    if r["bea_code"] == EMP_SVCS_DETAIL_CODE), None),
        }
        _record_run("bea_sut_naics_concordance", uri, SOURCE_URL, len(rows), built, cov,
                    status, error_text, started_at, completed_at)
        print(f"BEA_SUT_NAICS_CONCORDANCE SUMMARY: {cov} status={status}", flush=True)
    return {"dataset": "bea_sut_naics_concordance", "rows": len(rows), "indexes": built,
            "coverage": cov, "status": status}


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM 4 — derived contingent-labor intake (industry × year)
# ═══════════════════════════════════════════════════════════════════════════════════
def _derived_schema():
    import pyarrow as pa

    return pa.schema([
        ("industry_code", pa.string()), ("industry_name", pa.string()), ("year", pa.int32()),
        ("grain", pa.string()), ("emp_svcs_commodity_code", pa.string()),
        ("emp_svcs_input_musd", pa.float64()), ("total_intermediate_musd", pa.float64()),
        ("total_output_musd", pa.float64()),
        ("intake_share_of_output", pa.float64()), ("intake_share_of_intermediate", pa.float64()),
        ("source", pa.string()), ("source_file", pa.string()), ("source_url", pa.string()),
        ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


def _derived_from_spec(spec: dict, emp_code: str, grain_label: str, ingested_at) -> list[dict]:
    """Pull only the three load-bearing rows per sheet (employment services, total intermediate
    inputs, total industry output) across the industry columns."""
    wb = _workbook(spec["member"])
    out: list[dict] = []
    try:
        for sheet in spec["sheets"]:
            ws = wb[sheet]
            cols, body = _sheet_header(ws, spec)
            ccol = spec["code_col"]
            wanted = {emp_code: None, TOTAL_INTERMEDIATE_ROW: None, TOTAL_OUTPUT_ROW: None}
            for r in body:
                rc = _code(r[ccol]) if ccol < len(r) else None
                if rc in wanted and wanted[rc] is None:
                    wanted[rc] = r
            missing = [k for k, v in wanted.items() if v is None]
            if missing:
                raise RuntimeError(f"{sheet}: required row(s) {missing} absent")
            emp_r, int_r, out_r = (wanted[emp_code], wanted[TOTAL_INTERMEDIATE_ROW],
                                   wanted[TOTAL_OUTPUT_ROW])
            for j, icode, iname, ckind in cols:
                if ckind != "industry":
                    continue
                emp = _cell(emp_r[j])[0] if j < len(emp_r) else None
                if not emp:  # no staffing purchase (blank, marker, or a published zero)
                    continue
                inter = _cell(int_r[j])[0] if j < len(int_r) else None
                output = _cell(out_r[j])[0] if j < len(out_r) else None
                out.append({
                    "industry_code": icode, "industry_name": iname, "year": int(sheet),
                    "grain": grain_label, "emp_svcs_commodity_code": emp_code,
                    "emp_svcs_input_musd": emp, "total_intermediate_musd": inter,
                    "total_output_musd": output,
                    "intake_share_of_output": (emp / output) if output else None,
                    "intake_share_of_intermediate": (emp / inter) if inter else None,
                    "source": f"{SOURCE_TAG}:{grain_label}", "source_file": spec["member"],
                    "source_url": SOURCE_URL, "ingested_at": ingested_at,
                })
    finally:
        wb.close()
    return out


def run_derived(*, storage_options: dict, uri: str = DERIVED_URI) -> dict:
    import pyarrow as pa

    started_at = dt.datetime.now(dt.timezone.utc)
    ingested_at = started_at
    status, error_text, built = "error", None, []
    rows: list[dict] = []
    cov: dict = {}
    try:
        detail = _derived_from_spec(DETAIL_SPEC, EMP_SVCS_DETAIL_CODE, "detail_561300", ingested_at)
        summary = _derived_from_spec(SUMMARY_SPEC, EMP_SVCS_SUMMARY_CODE, "summary_561_proxy",
                                     ingested_at)
        rows = detail + summary
        print(f"  [derived] detail={len(detail):,} summary={len(summary):,}", flush=True)
        _gate_derived(detail, summary)

        _write_lance(pa.Table.from_pylist(rows, schema=_derived_schema()), uri, storage_options)
        built = _build_indexes(uri, btree=["industry_code"],
                               bitmap=["grain", "year"], so=storage_options)
        status = "success"
    except Exception as exc:  # noqa: BLE001
        error_text = str(exc)
        print(f"FATAL bea_contingent_labor_intake: {exc}", flush=True)
        raise
    finally:
        completed_at = dt.datetime.now(dt.timezone.utc)

        def _shares(subset):
            return [r["intake_share_of_output"] for r in subset
                    if r["intake_share_of_output"] is not None]

        d = [r for r in rows if r["grain"] == "detail_561300"]
        s = [r for r in rows if r["grain"] == "summary_561_proxy"]
        ds, ss = _shares(d), _shares(s)
        top = sorted(d, key=lambda r: r["intake_share_of_output"] or 0, reverse=True)
        cov = {
            "rows": len(rows), "detail_rows": len(d), "summary_rows": len(s),
            "detail_rows_by_year": {y: sum(1 for r in d if r["year"] == y)
                                    for y in sorted({r["year"] for r in d})},
            "summary_years": len({r["year"] for r in s}),
            "detail_share_range": [round(min(ds), 6), round(max(ds), 6)] if ds else None,
            "summary_share_range": [round(min(ss), 6), round(max(ss), 6)] if ss else None,
            "top10_2017_detail": [
                (r["industry_code"], r["industry_name"], round(r["intake_share_of_output"], 4))
                for r in top if r["year"] == 2017][:10],
        }
        _record_run("bea_contingent_labor_intake", uri, SOURCE_URL, len(rows), built, cov,
                    status, error_text, started_at, completed_at)
        print(f"BEA_CONTINGENT_LABOR_INTAKE SUMMARY: {cov} status={status}", flush=True)
    return {"dataset": "bea_contingent_labor_intake", "rows": len(rows), "indexes": built,
            "coverage": cov, "status": status}


def _gate_derived(detail: list[dict], summary: list[dict]) -> None:
    """Directive §8.4 — fail closed before anything is written."""
    years = {r["year"] for r in detail}
    if years != {int(y) for y in GATE_DETAIL_YEARS}:
        raise RuntimeError(f"GATE derived: detail years {sorted(years)} != {sorted(GATE_DETAIL_YEARS)}")
    for subset, label, bound in ((detail, "detail_561300", GATE_DETAIL_SHARE_MAX),
                                 (summary, "summary_561_proxy", GATE_SUMMARY_SHARE_MAX)):
        bad = [r for r in subset if r["intake_share_of_output"] is not None
               and not 0 < r["intake_share_of_output"] <= bound]
        if bad:
            worst = max(bad, key=lambda r: r["intake_share_of_output"])
            raise RuntimeError(
                f"GATE derived {label}: {len(bad)} rows outside (0, {bound}] — worst "
                f"{worst['industry_code']} {worst['year']} = {worst['intake_share_of_output']:.4f}")
    # Staffing-heavy sanity: the top decile of 2017 detail intake must contain at least one
    # administrative/support, health, or transportation/warehousing industry.
    y2017 = sorted((r for r in detail if r["year"] == 2017),
                   key=lambda r: r["intake_share_of_output"] or 0, reverse=True)
    decile = y2017[:max(1, len(y2017) // 10)]
    hits = [r["industry_name"] for r in decile
            if r["industry_name"] and GATE_STAFFING_HEAVY_RX.search(r["industry_name"])]
    if not hits:
        raise RuntimeError("GATE derived: no staffing-heavy industry in the 2017 top decile")
    ds = [r["intake_share_of_output"] for r in detail if r["intake_share_of_output"] is not None]
    ss = [r["intake_share_of_output"] for r in summary if r["intake_share_of_output"] is not None]
    print(f"  GATE derived ✓ detail={len(detail)} share<=[{max(ds):.4f}] "
          f"summary={len(summary)} share<=[{max(ss):.4f}] "
          f"top-decile staffing-heavy hits={len(hits)}", flush=True)


# ═══════════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════════
_STREAMS = ("detail", "summary", "concordance", "derived")
_URIS = {"detail": DETAIL_URI, "summary": SUMMARY_URI,
         "concordance": CONCORDANCE_URI, "derived": DERIVED_URI}
_RUNNERS = {"detail": run_detail, "summary": run_summary,
            "concordance": run_concordance, "derived": run_derived}


def _smoke_uri(uri: str) -> str:
    return f"s3://{BUCKET}/active/_smoke_{uri.rstrip('/').rsplit('/', 1)[-1]}/"


def _cli() -> None:
    p = argparse.ArgumentParser(
        description="BEA IO Use tables (Supply-Use Framework) → Lance SoR.")
    p.add_argument("--stream", choices=[*_STREAMS, "all"], required=True)
    p.add_argument("--smoke", action="store_true",
                   help="write to throwaway _smoke_ URIs and delete them after.")
    a = p.parse_args()

    so = _storage_options()
    streams = list(_STREAMS) if a.stream == "all" else [a.stream]
    uris = {k: (_smoke_uri(v) if a.smoke else v) for k, v in _URIS.items()}

    results = []
    for s in streams:
        print(f"\n=== stream {s} -> {uris[s]} ===", flush=True)
        results.append(_RUNNERS[s](storage_options=so, uri=uris[s]))

    if a.smoke:
        for s in streams:
            _delete_prefix(uris[s])

    print("\n=== BEA IO use ingest summary ===", flush=True)
    total = 0
    for r in results:
        total += r["rows"]
        print(f"  {r['dataset']:32s} rows={r['rows']:>9,}  idx={len(r['indexes']):>2}  "
              f"status={r['status']}", flush=True)
    print(f"  {'TOTAL':32s} rows={total:>9,}", flush=True)


if __name__ == "__main__":
    _cli()
