"""CBO Key Budget & Economic Data — landing corpus → Lance SoR (keyless, landing-drop route).

cbo.gov hard-blocks automated clients (Akamai edge 403; see the CBO directive §2.1). The
operator hand-drops the publisher's workbooks into s3://data-sink/landing/cbo/<product>/ and
this module melts them — NO cbo.gov fetch, NO GovInfo, NO api.data.gov key. (For future
automated refreshes the keyless route is CBO's public GitHub, not the keyed GovInfo API.)

CORPUS: 13 products (one landing subfolder each) × deep vintage history, ~347 files / ~159 MB,
formats .xlsx / legacy .xls / .zip (inner .xls/.xlsx). Filenames encode the vintage:
``{id}-{YYYY}-{MM}-{name}.{ext}``.

MELT (lossless cell-grain + navigation tags): each workbook is a "Contents" sheet + data tables
with banner rows then a row-label × year-column grid; units are encoded in the sheet name
((GDP)=% of GDP vs $B) — so we do NOT normalize units, we keep the raw cell + full context. One
landed row per non-empty cell, tagged with:
  - product / vintage_year / vintage_month / source_file / sheet
  - row_num, col_num, row_label (first text cell on the row), col_year (when the column's header
    is a 4-digit year), value_str (verbatim), value_num (coerced)
  - is_projection = (col_year > vintage_year) — CBO baselines put actuals then projections in
    adjacent year columns; anything past the publication vintage is a projection.
Filtering (row_label IS NOT NULL AND col_year IS NOT NULL) yields a clean
(product, vintage, sheet, measure, year, value) time series for reports / data-viz. Raw cells
remain for anything the tagging misses. NO LLM — deterministic openpyxl/xlrd parse only.

    doppler run -p core-x -c prd -- uv run --with pylance --with pyarrow --with openpyxl \\
      --with xlrd --with boto3 --with 'psycopg[binary]' \\
      python -m pipelines.reference.cbo_landing_ingest --stream all   # (--smoke for 2 files/product)
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import os
import re
import uuid
import zipfile

os.environ.setdefault("AWS_RESPONSE_CHECKSUM_VALIDATION", "when_required")
os.environ.setdefault("AWS_REQUEST_CHECKSUM_CALCULATION", "when_required")

from pipelines.bls.ingest import (  # noqa: E402
    DATA_STORAGE_VERSION, MAX_BYTES_PER_FILE, MAX_ROWS_PER_FILE, _build_indexes,
    _s3_client, _storage_options,
)

BUCKET = "data-sink"
URI = f"s3://{BUCKET}/active/cbo_key_budget_economic_data/"
# Melt reads the DURABLE raw archive under active/cbo_raw/ (the operator's as-is copy of the
# landing drops) so a re-run does not depend on landing/ (transport-only, may be cleared).
RAW_PREFIX = "active/cbo_raw/"
PRODUCTS = [
    "10-year-budget-projections", "long-term-budget-projections", "historical-budget-data",
    "10-year-trust-fund-projections", "revenue-projections-by-category", "spending-projections",
    "estimates-of-automatic-stabilizers", "tax-parameters", "economic-projections",
    "historical-data-and-economic-projections", "potential-gdp-and-underlying-inputs",
    "long-term-economic-projections", "demographic-projections",
]
VALUE_STR_CAP = 300


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    if not s or s in {"-", "--", "n.a.", "N/A", "NA", "*", "...", "…"}:
        return None
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _year_of(v):
    """Return the 4-digit year (1900–2100) iff the cell is essentially JUST a year header
    ("2026", "FY2026", "2026a") — not an embedded year in text, not a 21xx code."""
    if isinstance(v, (int, float)) and float(v).is_integer():
        y = int(v)
        return y if 1900 <= y <= 2100 else None
    s = _s(v)
    if s and re.fullmatch(r"(?:FY\s*)?(?:19|20)\d{2}[a-z]?", s.strip(), re.I):
        return int(re.search(r"(?:19|20)\d{2}", s).group(0))
    return None


def _parse_vintage(filename):
    # vintage is the {YYYY}-{MM} group AFTER the CBO product-id prefix; anchor on a real
    # 19xx/20xx year + valid month so a 5-digit product id ("55022-2000") isn't mis-parsed.
    m = re.search(r"((?:19|20)\d{2})-(0[1-9]|1[0-2])", filename)
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def _iter_sheets(raw, filename):
    """Yield (sheet_name, rows) for .xlsx (openpyxl) or legacy .xls (xlrd)."""
    low = filename.lower()
    if low.endswith(".xlsx") or low.endswith(".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            for sn in wb.sheetnames:
                yield sn, [tuple(r) for r in wb[sn].iter_rows(values_only=True)]
        finally:
            wb.close()
    else:
        import xlrd
        wb = xlrd.open_workbook(file_contents=raw)
        for sh in wb.sheets():
            yield sh.name, [tuple(sh.row_values(i)) for i in range(sh.nrows)]


def _melt_rows(rows, sheet, product, vy, vm, source_file, ingested_at):
    """Cell-grain melt of one sheet/CSV: one row per non-empty cell + col_year / row_label /
    is_projection tags. Year-header row = the row (in the first 20) with the most year cells."""
    out = []
    year_map, best = {}, 0
    for r in rows[:20]:
        m = {j: _year_of(c) for j, c in enumerate(r)}
        m = {j: y for j, y in m.items() if y is not None}
        if len(m) > best and len(m) >= 3:
            best, year_map = len(m), m
    for ri, r in enumerate(rows):
        if not r:
            continue
        row_label = next((_s(c) for c in r if _s(c) and _num(c) is None), None)
        for ci, cell in enumerate(r):
            vs = _s(cell)
            if vs is None:
                continue
            cy = year_map.get(ci)
            out.append({
                "product": product, "vintage_year": vy, "vintage_month": vm,
                "source_file": source_file, "sheet": sheet,
                "row_num": ri + 1, "col_num": ci + 1, "row_label": row_label, "col_year": cy,
                "is_projection": (cy is not None and vy is not None and cy > vy),
                "value_str": vs[:VALUE_STR_CAP], "value_num": _num(cell),
                "ingested_at": ingested_at,
            })
    return out


def _melt_workbook(raw, filename, product, source_file, ingested_at, *, vy=None, vm=None):
    """Cell-grain melt over every sheet of an .xls/.xlsx workbook."""
    if vy is None:
        vy, vm = _parse_vintage(filename)
    out = []
    for sheet, rows in _iter_sheets(raw, filename):
        out += _melt_rows(rows, sheet, product, vy, vm, source_file, ingested_at)
    return out


def _melt_csv(raw, filename, product, source_file, ingested_at, *, vy, vm):
    """Cell-grain melt of a CSV (CBO zips carry economic-data CSVs, not workbooks)."""
    import csv
    text = raw.decode("utf-8", "replace")
    rows = [tuple(r) for r in csv.reader(io.StringIO(text))]
    return _melt_rows(rows, filename, product, vy, vm, source_file, ingested_at)


def _schema():
    import pyarrow as pa
    return pa.schema([
        ("product", pa.string()), ("vintage_year", pa.int32()), ("vintage_month", pa.int32()),
        ("source_file", pa.string()), ("sheet", pa.string()),
        ("row_num", pa.int32()), ("col_num", pa.int32()), ("row_label", pa.string()),
        ("col_year", pa.int32()), ("is_projection", pa.bool_()),
        ("value_str", pa.string()), ("value_num", pa.float64()),
        ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


def _list_product_files(s3, product):
    out, token = [], None
    while True:
        kw = {"Bucket": BUCKET, "Prefix": f"{RAW_PREFIX}{product}/"}
        if token:
            kw["ContinuationToken"] = token
        r = s3.list_objects_v2(**kw)
        out += [o["Key"] for o in r.get("Contents", [])]
        token = r.get("NextContinuationToken")
        if not token:
            break
    return [k for k in out if not k.endswith("/.keep") and not k.endswith("/")
            and k.lower().rsplit(".", 1)[-1] in ("xls", "xlsx", "xlsm", "zip")]


def _melt_file(raw, key, product, ingested_at):
    """Melt one landing object: .xls/.xlsx workbook, or a .zip of workbooks/CSVs (vintage taken
    from the outer file/zip name; inner READMEs and other non-data files are skipped)."""
    base = key.rsplit("/", 1)[-1]
    vy, vm = _parse_vintage(base)
    if base.lower().endswith(".zip"):
        rows = []
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            for name in zf.namelist():
                if name.endswith("/") or name.startswith("__"):
                    continue
                ext = name.lower().rsplit(".", 1)[-1] if "." in name.rsplit("/", 1)[-1] else ""
                inner_src = f"{base}::{name.rsplit('/', 1)[-1]}"
                try:
                    if ext in ("xls", "xlsx", "xlsm"):
                        rows += _melt_workbook(zf.read(name), name, product, inner_src, ingested_at, vy=vy, vm=vm)
                    elif ext == "csv":
                        rows += _melt_csv(zf.read(name), name.rsplit("/", 1)[-1], product, inner_src,
                                          ingested_at, vy=vy, vm=vm)
                except Exception as exc:  # noqa: BLE001 — one bad inner file never kills the zip
                    print(f"    WARN inner {name}: {exc}", flush=True)
        return rows
    return _melt_workbook(raw, base, product, base, ingested_at, vy=vy, vm=vm)


def _record_run(run_id, rows, files, failed, status, notes):
    dsn = os.environ.get("HQX_DB_URL_POOLED")
    if not dsn:
        return
    try:
        import psycopg
        with psycopg.connect(dsn, autocommit=True) as c, c.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS ops.federal_appropriations_ingest_runs (
                    id bigint GENERATED ALWAYS AS IDENTITY PRIMARY KEY, run_id uuid NOT NULL,
                    stream text NOT NULL, resolved_url text, source_bytes bigint, rows_written bigint,
                    datasets jsonb, started_at timestamptz, finished_at timestamptz,
                    status text NOT NULL CHECK (status IN ('running','completed','failed')),
                    disposition text, notes text, recorded_at timestamptz NOT NULL DEFAULT now());""")
            cur.execute("""
                INSERT INTO ops.federal_appropriations_ingest_runs
                    (run_id, stream, resolved_url, source_bytes, rows_written, datasets,
                     started_at, finished_at, status, disposition, notes)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (str(run_id), "cbo_key_budget_economic_data", "active/cbo_raw/", None, rows,
                 json.dumps({"cbo_key_budget_economic_data": rows, "files": files}),
                 dt.datetime.now(dt.timezone.utc), dt.datetime.now(dt.timezone.utc), status,
                 "partial" if failed else "ok", notes))
    except Exception as exc:  # noqa: BLE001
        print(f"WARN: ledger write failed: {exc}", flush=True)


def run(products, *, smoke: bool):
    import lance
    import pyarrow as pa

    so = _storage_options()
    s3 = _s3_client()
    uri = URI.replace("/active/", "/smoke/") if smoke else URI
    ingested_at = dt.datetime.now(dt.timezone.utc)
    run_id = uuid.uuid4()
    total_rows = total_files = 0
    failed = []
    first = True
    per_product = {}
    for product in products:
        files = _list_product_files(s3, product)
        if smoke:
            files = files[:2]
        rows = []
        for key in files:
            try:
                raw = s3.get_object(Bucket=BUCKET, Key=key)["Body"].read()
                rows += _melt_file(raw, key, product, ingested_at)
                total_files += 1
            except Exception as exc:  # noqa: BLE001 — one bad file is logged, never fatal
                failed.append({"key": key, "error": str(exc)[:200]})
                print(f"  WARN {key}: {exc}", flush=True)
        if rows:
            tbl = pa.Table.from_pylist(rows, schema=_schema())
            lance.write_dataset(tbl, uri, mode=("overwrite" if first else "append"),
                                data_storage_version=DATA_STORAGE_VERSION,
                                max_rows_per_file=MAX_ROWS_PER_FILE, max_bytes_per_file=MAX_BYTES_PER_FILE,
                                storage_options=so)
            first = False
            per_product[product] = len(rows)
            total_rows += len(rows)
        print(f"  {product:44s} files={len(files):>3} rows={len(rows):>9,}", flush=True)
        del rows

    built = _build_indexes(uri, btree=["product", "vintage_year", "col_year"], bitmap=[], so=so)
    status = "completed" if not failed else "completed"
    notes = f"products={len(per_product)}; files={total_files}; per_product={json.dumps(per_product)}; failed={len(failed)}"
    if not smoke:
        _record_run(run_id, total_rows, total_files, failed, status, notes[:8000])
    print(f"\ncbo_key_budget_economic_data -> {uri}")
    print(f"  files={total_files}  rows={total_rows:,}  indexes={built}  failed={len(failed)}")
    if failed:
        print("  failures:", json.dumps(failed[:10]))
    return {"rows": total_rows, "files": total_files, "failed": failed, "uri": uri}


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--stream", default="all", help="'all' or a product folder name")
    ap.add_argument("--smoke", action="store_true", help="first 2 files/product → smoke/ URI")
    a = ap.parse_args()
    products = PRODUCTS if a.stream == "all" else [a.stream]
    run(products, smoke=a.smoke)


if __name__ == "__main__":
    main()
