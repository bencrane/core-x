"""Industry Cost-Structure Batch — KLEMS + ACES + IRS SOI + BEA Fixed Assets + QCEW + BDS + MFP/TFP → Gen-3 Lance SoR.

Directive: hq/directives/2026-07-23-industry-cost-structure-batch-ingest.md. Closes the
"where does the industry dollar go" decomposition beyond the labor-share stack (SUSB × ECEC ×
BEA value-added, PR #1118–#1120): the K/L/E/M/S recipe, the equipment capex flow and capital
stock, filed income-statement cost lines, firm demographics, and county-grain wages.

Eight streams, keyless public statistical sources, Pattern A (ephemeral download → parse →
lance.write_dataset → s3://data-sink/active/<name>/). No RisingWave, no Postgres source
tables, no catalog layer. Raw stays lossless (suppression sentinels land verbatim; cleaned
numerics are ADDITIONAL columns). ASM was dropped at Stage 1: no keyless expense-detail
files exist on www2.census.gov (only a robotics table) — gap documented in the run record.

  --stream klems   → active/bea_bls_klems/            BEA-BLS integrated production account
                     1997–2024 (ONE xlsx; 31 measure sheets melted long; industry rows are
                     joined to production-account codes + 2017 NAICS via the workbook's own
                     concordance sheet).
  --stream bea_fa  → active/bea_fixed_assets_detail/  BEA detailed fixed assets: per-industry
                     sheets (75), asset × year matrices; stk1 (current-cost net stock,
                     1925–2024) + inv1 (investment, 1901–2024) discriminated by `measure`.
  --stream aces    → active/census_aces_capex/        Census ACES tables 1998–2022, EVERY
                     xls/xlsx in each year dir (incl. OLD- superseded variants — no data left
                     behind), landed at CELL grain (sheet, row, col, row_label, value) —
                     presentation workbooks vary too much for a schema'd melt to be honest.
  --stream soi     → active/irs_soi_corp_industry/    IRS SOI corporation complete-report
                     tables {yy}co{NN}ccr.xlsx, tax years 2014–2022, tables 01–26 probed per
                     year; CELL grain (same rationale as aces).
  --stream bds     → active/census_bds_firm_dynamics/ Census BDS 1978–2023: EVERY variant CSV
                     in the release dir (economy/sector/NAICS/age/size + state/county/MSA/metro
                     crosses; 133 files) union'd by name via DuckDB, all-varchar (suppression
                     flags (D)/(S)/(X) preserved), `variant` from filename.
  --stream qcew    → active/bls_qcew_annual/          QCEW annual-average singlefiles
                     2001–2024, lossless 38-col layout, typed casts + varchar passthrough for
                     any unknown col. Each downloaded zip is ALSO parked verbatim under
                     s3://data-sink/landing/bls/qcew/ (operator convention); if the zip is
                     already in landing it is used instead of the upstream fetch.
  --stream mfp     → active/bls_mfp_major_sector/ + active/bls_tfp_detailed_industries/
                     BLS TFP from the operator-landed catalog (bls.gov hard-403s agents;
                     s3://data-sink/landing/bls/productivity/, 27 files, 2026-07-23): the
                     MachineReadable sheets of major-sectors (+historical) and the two
                     detailed-industry workbooks.
  --stream prod_catalog → active/bls_productivity_tables/ (MachineReadable union) +
                     active/bls_productivity_cells/ (cell-grain fallback for workbooks
                     without a MachineReadable sheet). The REST of the landed catalog —
                     labor productivity, hours, capital details, rental prices, dispersion,
                     contributions-to-growth, government, historical/SIC, experimental.

Ledger: ops.industry_cost_structure_runs (HQX_DB_URL_POOLED; create-if-not-exists inline;
an audit-write failure warns, never masks a good load).

    doppler run -p core-x -c prd -- uv run --with pylance --with pyarrow --with duckdb \\
      --with openpyxl --with xlrd --with requests --with boto3 --with 'psycopg[binary]' \\
      python -m pipelines.reference.industry_cost_structure_ingest --stream all --smoke
    ... --stream all                                                                # full
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import os
import re
import tempfile
import time
import zipfile

# Cyberduck multipart uploads to R2 carry composite checksums that newer AWS SDKs
# reject as full-object mismatches — validate only when explicitly requested.
os.environ.setdefault("AWS_RESPONSE_CHECKSUM_VALIDATION", "when_required")
os.environ.setdefault("AWS_REQUEST_CHECKSUM_CALCULATION", "when_required")

# Reuse the fleet R2/index plumbing verbatim (do not reimplement).
from pipelines.bls.ingest import (  # noqa: E402
    DATA_STORAGE_VERSION,
    MAX_BYTES_PER_FILE,
    MAX_ROWS_PER_FILE,
    _build_indexes,
    _s3_client,
    _storage_options,
)

BUCKET = "data-sink"
SOURCE_TAG = "industry_cost_structure_ingest"

URIS = {
    "bea_bls_klems": f"s3://{BUCKET}/active/bea_bls_klems/",
    "bea_fixed_assets_detail": f"s3://{BUCKET}/active/bea_fixed_assets_detail/",
    "census_aces_capex": f"s3://{BUCKET}/active/census_aces_capex/",
    "irs_soi_corp_industry": f"s3://{BUCKET}/active/irs_soi_corp_industry/",
    "census_bds_firm_dynamics": f"s3://{BUCKET}/active/census_bds_firm_dynamics/",
    "bls_qcew_annual": f"s3://{BUCKET}/active/bls_qcew_annual/",
    "bls_mfp_major_sector": f"s3://{BUCKET}/active/bls_mfp_major_sector/",
    "bls_tfp_detailed_industries": f"s3://{BUCKET}/active/bls_tfp_detailed_industries/",
    "bls_productivity_tables": f"s3://{BUCKET}/active/bls_productivity_tables/",
    "bls_productivity_cells": f"s3://{BUCKET}/active/bls_productivity_cells/",
}

KLEMS_URL = ("https://www.bea.gov/sites/default/files/2026-04/"
             "BEA-BLS-industry-level-production-account-1997-2024.xlsx")
FA_URLS = {
    "current_cost_net_stock": "https://apps.bea.gov/national/FA2004/Details/xls/detailnonres_stk1.xlsx",
    "investment": "https://apps.bea.gov/national/FA2004/Details/xls/detailnonres_inv1.xlsx",
}
ACES_DIR = "https://www2.census.gov/programs-surveys/aces/tables/{year}/"
ACES_YEARS = list(range(1998, 2023))
SOI_URL = "https://www.irs.gov/pub/irs-soi/{yy}co{nn}ccr.xlsx"
SOI_YEARS = list(range(2014, 2023))
SOI_TABLES = [f"{n:02d}" for n in range(1, 27)]
BDS_DIR = "https://www2.census.gov/programs-surveys/bds/tables/time-series/2023/"
BDS_LANDING = "landing/census/bds/"
QCEW_URL = "https://data.bls.gov/cew/data/files/{year}/csv/{year}_annual_singlefile.zip"
QCEW_YEARS = list(range(2001, 2025))
QCEW_LANDING = "landing/bls/qcew/"
PROD_LANDING = "landing/bls/productivity/"

# landing → stream routing for the productivity catalog
MFP_MAJOR_FILES = [
    "total-factor-productivity-major-sectors.xlsx",
    "total-factor-productivity-major-sectors-historical.xlsx",
]
TFP_DETAILED_FILES = [
    "major-industry-total-factor-productivity-klems.xlsx",
    "total-factor-productivity-manufacturing-and-transportation-detailed-industries.xlsx",
]

_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/126 Safari/537.36")


# ── HTTP ────────────────────────────────────────────────────────────────────────────
def _download(url: str, *, timeout: int = 600, ok404: bool = False) -> bytes | None:
    import requests

    last = None
    for attempt in range(6):
        try:
            resp = requests.get(url, headers={"User-Agent": _UA}, timeout=timeout)
        except requests.RequestException as exc:  # noqa: PERF203
            last = exc
            time.sleep(min(30, 2 ** attempt))
            continue
        if resp.status_code == 200:
            return resp.content
        if resp.status_code == 404 and ok404:
            return None
        if resp.status_code in (429, 500, 502, 503, 504):
            last = RuntimeError(f"HTTP {resp.status_code}")
            time.sleep(min(60, 2 ** attempt))
            continue
        raise RuntimeError(f"GET {url} HTTP {resp.status_code}: {(resp.text or '')[:200]}")
    raise RuntimeError(f"GET {url}: retries exhausted ({last}).")


def _dir_files(url: str, exts: tuple[str, ...]) -> list[str]:
    """Names of data files in a www2.census.gov autoindex page."""
    raw = _download(url, ok404=True)
    if raw is None:
        return []
    hrefs = re.findall(r'href="([^"?/][^"]*)"', raw.decode("utf-8", "replace"))
    return sorted({h for h in hrefs if h.lower().endswith(exts)})


# ── coercion ────────────────────────────────────────────────────────────────────────
def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if not s or s in {"(D)", "(S)", "(X)", "(NA)", "(Z)", "...", "…", "*", "**", "-", "--",
                      "N.A.", "NA", "n.a."}:
        return None
    if s.startswith("(") and s.endswith(")"):  # accounting negatives
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _year(v) -> int | None:
    s = _s(v)
    if s is None:
        return None
    s = s.split(".")[0]
    return int(s) if s.isdigit() and len(s) == 4 else None


# ── ledger ──────────────────────────────────────────────────────────────────────────
def _record_run(dataset: str, uri: str, source: str, rows: int, built: list,
                coverage: dict, status: str, error: str | None, started_at, completed_at) -> None:
    dsn = os.environ.get("HQX_DB_URL_POOLED")
    if not dsn:
        print("WARN: HQX_DB_URL_POOLED not set; skipping ledger write.", flush=True)
        return
    try:
        import psycopg

        with psycopg.connect(dsn) as conn, conn.cursor() as cur:
            cur.execute("CREATE SCHEMA IF NOT EXISTS ops;")
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS ops.industry_cost_structure_runs (
                    id bigint GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                    dataset text NOT NULL, dataset_uri text, source_url text,
                    rows_processed bigint, indexes_built text[], coverage jsonb,
                    status text NOT NULL, error text,
                    started_at timestamptz, completed_at timestamptz,
                    recorded_at timestamptz NOT NULL DEFAULT now());
                """
            )
            cur.execute(
                """
                INSERT INTO ops.industry_cost_structure_runs
                    (dataset, dataset_uri, source_url, rows_processed, indexes_built,
                     coverage, status, error, started_at, completed_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (dataset, uri, source, rows, built, json.dumps(coverage), status, error,
                 started_at, completed_at),
            )
            conn.commit()
    except Exception as exc:  # noqa: BLE001 — audit must never mask a good load
        print(f"WARN: ledger write failed: {exc}", flush=True)


def _write_lance(table, uri: str, so: dict, *, mode: str = "overwrite") -> None:
    import lance

    lance.write_dataset(table, uri, mode=mode,
                        data_storage_version=DATA_STORAGE_VERSION,
                        max_rows_per_file=MAX_ROWS_PER_FILE, max_bytes_per_file=MAX_BYTES_PER_FILE,
                        storage_options=so)


def _finish(dataset: str, uri: str, source: str, rows: int, built: list, cov: dict,
            status: str, err: str | None, t0) -> dict:
    _record_run(dataset, uri, source, rows, built, cov, status, err,
                t0, dt.datetime.now(dt.timezone.utc))
    print(f"{dataset} SUMMARY: rows={rows} cov={json.dumps(cov)[:400]} status={status}", flush=True)
    return {"dataset": dataset, "rows": rows, "coverage": cov, "status": status}


# ── workbook loaders (xlsx via openpyxl, xls via xlrd) — yield rows of values ──────
def _iter_sheets(raw: bytes, filename: str):
    """Yield (sheet_name, row_iterator) for xlsx OR legacy xls."""
    if filename.lower().endswith(".xlsx"):
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            for sn in wb.sheetnames:
                yield sn, wb[sn].iter_rows(values_only=True)
        finally:
            wb.close()
    else:
        import xlrd

        wb = xlrd.open_workbook(file_contents=raw)
        for sh in wb.sheets():
            def _rows(sh=sh):
                for i in range(sh.nrows):
                    yield tuple(sh.row_values(i))
            yield sh.name, _rows()


def _cell_rows(raw: bytes, filename: str, ref_year: int, ingested_at) -> list[dict]:
    """Lossless CELL-grain melt: one row per non-empty cell, with the row's first-column
    label carried for readability. The honest landing for presentation workbooks."""
    out: list[dict] = []
    for sheet, rows in _iter_sheets(raw, filename):
        for r_idx, row in enumerate(rows, start=1):
            label = _s(row[0]) if row else None
            for c_idx, cell in enumerate(row, start=1):
                s = _s(cell)
                if s is None:
                    continue
                out.append({
                    "source_file": filename, "ref_year": ref_year, "sheet": sheet,
                    "row_num": r_idx, "col_num": c_idx, "row_label": label,
                    "value_str": s[:4000], "value_num": _num(cell),
                    "source": SOURCE_TAG, "ingested_at": ingested_at,
                })
    return out


def _cells_schema():
    import pyarrow as pa
    return pa.schema([
        ("source_file", pa.string()), ("ref_year", pa.int32()), ("sheet", pa.string()),
        ("row_num", pa.int32()), ("col_num", pa.int32()), ("row_label", pa.string()),
        ("value_str", pa.string()), ("value_num", pa.float64()),
        ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


# ── R2 landing helpers ──────────────────────────────────────────────────────────────
def _landing_list(prefix: str) -> list[str]:
    s3 = _s3_client()
    out, token = [], None
    while True:
        kw = {"Bucket": BUCKET, "Prefix": prefix}
        if token:
            kw["ContinuationToken"] = token
        resp = s3.list_objects_v2(**kw)
        out += [o["Key"] for o in resp.get("Contents", [])]
        token = resp.get("NextContinuationToken")
        if not token:
            break
    return [k for k in out if not k.endswith("/.keep") and not k.endswith("/")]


def _landing_get(key: str) -> bytes:
    s3 = _s3_client()
    return s3.get_object(Bucket=BUCKET, Key=key)["Body"].read()


def _landing_put(key: str, body: bytes) -> None:
    _s3_client().put_object(Bucket=BUCKET, Key=key, Body=body)


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM klems — BEA-BLS integrated production account (K/L/E/M/S + TFP)
# ═══════════════════════════════════════════════════════════════════════════════════
def run_klems(*, so: dict, uri: str, smoke: bool) -> dict:
    import openpyxl
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    built: list = []
    status, err = "error", None
    try:
        raw = _download(KLEMS_URL)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        # concordance: description → (production code, 2017 NAICS)
        conc: dict[str, tuple[str | None, str | None]] = {}
        for r in wb["NAICS codes"].iter_rows(min_row=4, values_only=True):
            desc, code, naics = _s(r[0]), _s(r[1]), _s(r[2] if len(r) > 2 else None)
            if desc and code:
                conc[desc.lower()] = (code, naics)
        data_sheets = [sn for sn in wb.sheetnames if sn not in ("Readme (Integrated TFP)", "NAICS codes")]
        if smoke:
            data_sheets = data_sheets[:2]
        for sn in data_sheets:
            ws = wb[sn]
            it = ws.iter_rows(values_only=True)
            title = _s(next(it, (None,))[0])
            header = next(it, ())
            ycols = [(j, _year(c)) for j, c in enumerate(header) if _year(c)]
            for r in it:
                desc = _s(r[0]) if r else None
                if desc is None:
                    continue
                code, naics = conc.get(desc.lower(), (None, None))
                for j, yr in ycols:
                    if j >= len(r):
                        continue
                    v = _num(r[j])
                    vs = _s(r[j])
                    if vs is None:
                        continue
                    rows.append({
                        "sheet": sn, "measure_title": title, "industry_desc": desc,
                        "production_code": code, "naics_2017": naics, "year": yr,
                        "value_num": v, "value_str": vs,
                        "source": f"{SOURCE_TAG}:bea_klems", "ingested_at": t0,
                    })
        wb.close()
        schema = pa.schema([
            ("sheet", pa.string()), ("measure_title", pa.string()), ("industry_desc", pa.string()),
            ("production_code", pa.string()), ("naics_2017", pa.string()), ("year", pa.int32()),
            ("value_num", pa.float64()), ("value_str", pa.string()),
            ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
        ])
        tbl = pa.Table.from_pylist(rows, schema=schema)
        _write_lance(tbl, uri, so)
        built = _build_indexes(uri, btree=["production_code", "year"], bitmap=["sheet"], so=so)
        # gates
        years = {r["year"] for r in rows}
        inds = {r["industry_desc"] for r in rows}
        keys = {(r["sheet"], r["industry_desc"], r["year"]) for r in rows}
        if not smoke:
            assert min(years) == 1997 and max(years) == 2024, f"klems year span {min(years)}–{max(years)}"
            # 63 account industries + sector/aggregate rows on some sheets → ~124 labels
            assert 50 <= len(inds) <= 150, f"klems industry count {len(inds)}"
            assert len(keys) == len(rows), "klems duplicate (sheet, industry, year)"
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        cov = {"rows": len(rows), "sheets": len({r['sheet'] for r in rows}),
               "industries": len({r['industry_desc'] for r in rows}),
               "years": [min({r['year'] for r in rows}, default=None),
                          max({r['year'] for r in rows}, default=None)]}
        res = _finish("bea_bls_klems", uri, KLEMS_URL, len(rows), built, cov, status, err, t0)
    return res


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM bea_fa — BEA detailed fixed assets (stock + investment, industry × asset)
# ═══════════════════════════════════════════════════════════════════════════════════
def run_bea_fa(*, so: dict, uri: str, smoke: bool) -> dict:
    import openpyxl
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    built: list = []
    status, err = "error", None
    try:
        for measure, url in FA_URLS.items():
            raw = _download(url)
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheets = [sn for sn in wb.sheetnames if sn != "readme"]
            if smoke:
                sheets = sheets[:3]
            for sn in sheets:
                ws = wb[sn]
                it = ws.iter_rows(values_only=True)
                first = next(it, ())
                ind_name = _s(first[0]) if first else None
                if ind_name and " - " in ind_name:
                    ind_name = ind_name.split(" - ", 1)[1].replace("\xa0", " ").strip()
                ycols: list[tuple[int, int]] = []
                for r in it:  # find the 'Asset Codes' header row
                    if r and _s(r[0]) == "Asset Codes":
                        ycols = [(j, _year(c)) for j, c in enumerate(r) if _year(c)]
                        break
                if not ycols:
                    print(f"WARN bea_fa: no header row in {measure}/{sn}; skipped", flush=True)
                    continue
                for r in it:
                    code, name = _s(r[0]) if r else None, _s(r[1]) if len(r) > 1 else None
                    if name is None:
                        continue
                    for j, yr in ycols:
                        if j >= len(r):
                            continue
                        v = _num(r[j])
                        if v is None:
                            continue
                        rows.append({
                            "industry_code": sn, "industry_name": ind_name,
                            "asset_code": code, "asset_name": name, "measure": measure,
                            "year": yr, "value_musd": v,
                            "source": f"{SOURCE_TAG}:bea_fa", "ingested_at": t0,
                        })
            wb.close()
        schema = pa.schema([
            ("industry_code", pa.string()), ("industry_name", pa.string()),
            ("asset_code", pa.string()), ("asset_name", pa.string()), ("measure", pa.string()),
            ("year", pa.int32()), ("value_musd", pa.float64()),
            ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
        ])
        tbl = pa.Table.from_pylist(rows, schema=schema)
        _write_lance(tbl, uri, so)
        built = _build_indexes(uri, btree=["industry_code", "asset_code", "year"],
                               bitmap=["measure"], so=so)
        if not smoke:
            years = {r["year"] for r in rows}
            names = " | ".join({r["asset_name"] for r in rows}).lower()
            assert 100_000 <= len(rows) <= 2_000_000, f"bea_fa row count {len(rows)}"
            assert min(years) <= 1925 and max(years) >= 2023, f"bea_fa year span {min(years)}–{max(years)}"
            assert {r["measure"] for r in rows} == set(FA_URLS), "bea_fa missing a measure"
            assert "construction machinery" in names and "trucks" in names, "bea_fa key assets missing"
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        cov = {"rows": len(rows), "industries": len({r['industry_code'] for r in rows}),
               "assets": len({r['asset_code'] for r in rows}),
               "measures": sorted({r['measure'] for r in rows})}
        res = _finish("bea_fixed_assets_detail", uri, " + ".join(FA_URLS.values()),
                      len(rows), built, cov, status, err, t0)
    return res


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM aces — Census ACES capex tables (cell grain, every file, every year)
# ═══════════════════════════════════════════════════════════════════════════════════
def run_aces(*, so: dict, uri: str, smoke: bool) -> dict:
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    built: list = []
    landed_years: dict[int, int] = {}
    status, err = "error", None
    years = [2022] if smoke else ACES_YEARS
    try:
        for year in years:
            files = _dir_files(ACES_DIR.format(year=year), (".xls", ".xlsx"))
            if smoke:
                files = files[:3]
            n0 = len(rows)
            for fn in files:
                raw = _download(ACES_DIR.format(year=year) + fn, ok404=True)
                if raw is None:
                    continue
                try:
                    rows += _cell_rows(raw, fn, year, t0)
                except Exception as exc:  # noqa: BLE001 — one bad workbook never blocks the batch
                    print(f"WARN aces {year}/{fn}: {exc}", flush=True)
            landed_years[year] = len(rows) - n0
            print(f"aces {year}: {len(files)} files, {len(rows) - n0} cells", flush=True)
        tbl = pa.Table.from_pylist(rows, schema=_cells_schema())
        _write_lance(tbl, uri, so)
        built = _build_indexes(uri, btree=["ref_year"], bitmap=["source_file"], so=so)
        if not smoke:
            missing = [y for y, n in landed_years.items() if n == 0]
            assert not missing, f"aces years with zero cells: {missing}"
            numeric = sum(1 for r in rows if r["value_num"] is not None)
            assert numeric >= 200_000, f"aces numeric cells {numeric}"
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        cov = {"rows": len(rows), "years": {str(k): v for k, v in landed_years.items()},
               "files": len({r['source_file'] for r in rows})}
        res = _finish("census_aces_capex", uri, ACES_DIR, len(rows), built, cov, status, err, t0)
    return res


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM soi — IRS SOI corporation complete-report tables (cell grain)
# ═══════════════════════════════════════════════════════════════════════════════════
def run_soi(*, so: dict, uri: str, smoke: bool) -> dict:
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    built: list = []
    landed: dict[int, list[str]] = {}
    status, err = "error", None
    years = [2021] if smoke else SOI_YEARS
    tables = SOI_TABLES[:3] if smoke else SOI_TABLES
    try:
        for year in years:
            yy = f"{year % 100:02d}"
            landed[year] = []
            for nn in tables:
                raw = _download(SOI_URL.format(yy=yy, nn=nn), ok404=True)
                if raw is None:
                    continue
                fn = f"{yy}co{nn}ccr.xlsx"
                try:
                    rows += _cell_rows(raw, fn, year, t0)
                    landed[year].append(nn)
                except Exception as exc:  # noqa: BLE001
                    print(f"WARN soi {fn}: {exc}", flush=True)
            print(f"soi {year}: tables {landed[year]}", flush=True)
        tbl = pa.Table.from_pylist(rows, schema=_cells_schema())
        _write_lance(tbl, uri, so)
        built = _build_indexes(uri, btree=["ref_year"], bitmap=["source_file"], so=so)
        if not smoke:
            for year in years:
                assert len(landed[year]) >= 8, f"soi {year}: only {len(landed[year])} tables"
            deprec_years = {r["ref_year"] for r in rows
                            if r["value_num"] is None and "depreciation" in (r["value_str"] or "").lower()}
            assert set(years) <= deprec_years, f"soi: depreciation header missing in {set(years) - deprec_years}"
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        cov = {"rows": len(rows), "tables_by_year": {str(k): v for k, v in landed.items()}}
        res = _finish("irs_soi_corp_industry", uri, SOI_URL, len(rows), built, cov, status, err, t0)
    return res


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM bds — Census Business Dynamics Statistics (every variant, DuckDB union)
# ═══════════════════════════════════════════════════════════════════════════════════
def run_bds(*, so: dict, uri: str, smoke: bool) -> dict:
    import duckdb

    t0 = dt.datetime.now(dt.timezone.utc)
    built: list = []
    n_rows = 0
    n_files = 0
    status, err = "error", None
    cov: dict = {}
    try:
        files = _dir_files(BDS_DIR, (".csv",))
        files = [f for f in files if f.startswith("bds2023")]
        if smoke:
            files = [f for f in files if f in ("bds2023.csv", "bds2023_sec.csv", "bds2023_fa.csv")]
        n_files = len(files)
        # Resume-through-landing: every fetched CSV is parked verbatim in R2
        # (landing/census/bds/) the moment it lands, so a WAF block mid-crawl
        # costs only the missing tail — the next run never re-fetches what is
        # already secured (and never re-aggravates the WAF for it).
        parked = {os.path.basename(k) for k in _landing_list(BDS_LANDING)}
        with tempfile.TemporaryDirectory() as td:
            for fn in files:
                if fn in parked:
                    raw = _landing_get(f"{BDS_LANDING}{fn}")
                else:
                    # Census WAF intermittently serves a "Request Rejected" HTML
                    # page instead of the CSV — validate content, retry, back off.
                    raw = b""
                    for attempt in range(6):
                        raw = _download(BDS_DIR + fn)
                        if not raw.lstrip()[:1] == b"<":
                            break
                        print(f"WARN bds {fn}: WAF rejection page; retry {attempt + 1}", flush=True)
                        time.sleep(20 * (attempt + 1))
                    if raw.lstrip()[:1] == b"<":
                        raise RuntimeError(f"bds {fn}: Census WAF rejection persisted "
                                           f"({len(parked)} files already parked; re-run resumes)")
                    _landing_put(f"{BDS_LANDING}{fn}", raw)
                    parked.add(fn)
                    time.sleep(1.0)  # stay under the WAF's rate radar
                with open(os.path.join(td, fn), "wb") as fh:
                    fh.write(raw)
            con = duckdb.connect()
            rel = con.sql(
                f"""
                SELECT regexp_extract(parse_filename(filename), 'bds2023_?(.*)\\.csv', 1) AS variant,
                       * EXCLUDE (filename)
                FROM read_csv('{td}/*.csv', union_by_name=true, all_varchar=true,
                              header=true, filename=true,
                              strict_mode=false, null_padding=true)
                """
            )
            tbl = rel.fetch_arrow_table()  # variant '' = the economy-wide file
            _write_lance(tbl, uri, so)
            n_rows = tbl.num_rows
            # gates + coverage from the landed shape
            years = con.sql("SELECT min(TRY_CAST(year AS INT)), max(TRY_CAST(year AS INT)) "
                            f"FROM read_csv('{td}/bds2023.csv', all_varchar=true)").fetchone()
            econ_2023 = con.sql("SELECT TRY_CAST(firms AS BIGINT) FROM "
                                f"read_csv('{td}/bds2023.csv', all_varchar=true) "
                                "WHERE year='2023'").fetchone()
            cov = {"rows": n_rows, "files": n_files, "year_span": list(years),
                   "economy_firms_2023": econ_2023[0] if econ_2023 else None}
            if not smoke:
                assert years[0] == 1978 and years[1] == 2023, f"bds year span {years}"
                assert econ_2023 and 5_000_000 <= econ_2023[0] <= 8_000_000, f"bds economy firms {econ_2023}"
            con.close()
        built = _build_indexes(uri, btree=["year"], bitmap=["variant"], so=so)
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        res = _finish("census_bds_firm_dynamics", uri, BDS_DIR, n_rows, built, cov, status, err, t0)
    return res


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM qcew — BLS QCEW annual singlefiles (typed, lossless, landing-parked)
# ═══════════════════════════════════════════════════════════════════════════════════
_QCEW_TYPES = {
    "area_fips": "VARCHAR", "own_code": "VARCHAR", "industry_code": "VARCHAR",
    "agglvl_code": "VARCHAR", "size_code": "VARCHAR", "year": "INTEGER", "qtr": "VARCHAR",
    "disclosure_code": "VARCHAR", "annual_avg_estabs": "BIGINT", "annual_avg_emplvl": "BIGINT",
    "total_annual_wages": "BIGINT", "taxable_annual_wages": "BIGINT",
    "annual_contributions": "BIGINT", "annual_avg_wkly_wage": "BIGINT", "avg_annual_pay": "BIGINT",
    "lq_disclosure_code": "VARCHAR", "lq_annual_avg_estabs": "DOUBLE",
    "lq_annual_avg_emplvl": "DOUBLE", "lq_total_annual_wages": "DOUBLE",
    "lq_taxable_annual_wages": "DOUBLE", "lq_annual_contributions": "DOUBLE",
    "lq_annual_avg_wkly_wage": "DOUBLE", "lq_avg_annual_pay": "DOUBLE",
    "oty_disclosure_code": "VARCHAR",
    "oty_annual_avg_estabs_chg": "BIGINT", "oty_annual_avg_estabs_pct_chg": "DOUBLE",
    "oty_annual_avg_emplvl_chg": "BIGINT", "oty_annual_avg_emplvl_pct_chg": "DOUBLE",
    "oty_total_annual_wages_chg": "BIGINT", "oty_total_annual_wages_pct_chg": "DOUBLE",
    "oty_taxable_annual_wages_chg": "BIGINT", "oty_taxable_annual_wages_pct_chg": "DOUBLE",
    "oty_annual_contributions_chg": "BIGINT", "oty_annual_contributions_pct_chg": "DOUBLE",
    "oty_annual_avg_wkly_wage_chg": "BIGINT", "oty_annual_avg_wkly_wage_pct_chg": "DOUBLE",
    "oty_avg_annual_pay_chg": "BIGINT", "oty_avg_annual_pay_pct_chg": "DOUBLE",
}


def run_qcew(*, so: dict, uri: str, smoke: bool) -> dict:
    import duckdb

    t0 = dt.datetime.now(dt.timezone.utc)
    built: list = []
    total = 0
    per_year: dict[int, dict] = {}
    status, err = "error", None
    years = [2024] if smoke else QCEW_YEARS
    landing = set(_landing_list(QCEW_LANDING))
    try:
        first = True
        for year in years:
            key = f"{QCEW_LANDING}{year}_annual_singlefile.zip"
            route = "landing" if key in landing else "upstream"
            raw = _landing_get(key) if route == "landing" else _download(QCEW_URL.format(year=year))
            if route == "upstream" and not smoke:
                _landing_put(key, raw)  # park the verbatim zip (operator convention)
            with tempfile.TemporaryDirectory() as td:
                with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                    names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
                    assert len(names) == 1, f"qcew {year}: unexpected namelist {names}"
                    zf.extract(names[0], td)
                csv_path = os.path.join(td, names[0])
                con = duckdb.connect()
                cols = [r[0] for r in con.sql(
                    f"DESCRIBE SELECT * FROM read_csv('{csv_path}', all_varchar=true, header=true)"
                ).fetchall()]
                sel = ", ".join(
                    f'TRY_CAST("{c}" AS {_QCEW_TYPES[c]}) AS "{c}"' if c in _QCEW_TYPES
                    else f'"{c}"'
                    for c in cols
                )
                tbl = con.sql(
                    f"SELECT {sel} FROM read_csv('{csv_path}', all_varchar=true, header=true)"
                ).fetch_arrow_table()
                # national private all-industries total — agglvl varies by vintage, so
                # constrain only the identity columns and take the total row.
                nat = con.sql(
                    """SELECT max(total_annual_wages) FROM tbl
                       WHERE area_fips='US000' AND own_code='5' AND industry_code='10'"""
                ).fetchone()
                con.close()
            _write_lance(tbl, uri, so, mode=("overwrite" if first else "append"))
            first = False
            total += tbl.num_rows
            per_year[year] = {"rows": tbl.num_rows, "route": route,
                              "national_private_wages": nat[0] if nat else None}
            print(f"qcew {year}: {tbl.num_rows} rows via {route} "
                  f"(nat private wages: {nat[0] if nat else None})", flush=True)
            if not smoke:
                assert 2_500_000 <= tbl.num_rows <= 4_500_000, f"qcew {year} rows {tbl.num_rows}"
                assert nat and 2e12 <= nat[0] <= 14e12, f"qcew {year} national wages {nat}"
        built = _build_indexes(uri, btree=["area_fips", "industry_code", "year"],
                               bitmap=["own_code", "agglvl_code"], so=so)
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        cov = {"rows": total, "per_year": {str(k): v for k, v in per_year.items()}}
        res = _finish("bls_qcew_annual", uri, QCEW_URL, total, built, cov, status, err, t0)
    return res


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAMS mfp + prod_catalog — the operator-landed BLS productivity catalog
# ═══════════════════════════════════════════════════════════════════════════════════
_MR_MAP = {  # normalized MachineReadable header → landed column
    "naics": "naics", "sector": "sector", "industry": "industry", "digit": "digit",
    "basis": "basis", "measure": "measure", "units": "units", "year": "year",
    "value": "value", "state": "area", "region": "area", "area": "area",
}


def _mr_rows(raw: bytes, filename: str, ingested_at) -> list[dict] | None:
    """Rows from a workbook's MachineReadable sheet, or None if it has none."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        target = next((sn for sn in wb.sheetnames if sn.lower().replace(" ", "") == "machinereadable"), None)
        if target is None:
            return None
        it = wb[target].iter_rows(values_only=True)
        header = [(_s(c) or "").strip().lower() for c in next(it, ())]
        out = []
        for r in it:
            if not any(c is not None for c in r):
                continue
            rec = {"source_file": filename, "naics": None, "sector": None, "industry": None,
                   "digit": None, "basis": None, "measure": None, "units": None, "area": None,
                   "year": None, "value_num": None, "value_str": None, "extra": None,
                   "source": SOURCE_TAG, "ingested_at": ingested_at}
            extra = {}
            for h, c in zip(header, r):
                col = _MR_MAP.get(h)
                if col == "year":
                    rec["year"] = _year(c)
                elif col == "value":
                    rec["value_num"] = _num(c)
                    rec["value_str"] = _s(c)
                elif col:
                    rec[col] = _s(c)
                elif h:
                    v = _s(c)
                    if v is not None:
                        extra[h] = v
            if extra:
                rec["extra"] = json.dumps(extra)
            out.append(rec)
        return out
    finally:
        wb.close()


def _mr_schema():
    import pyarrow as pa
    return pa.schema([
        ("source_file", pa.string()), ("naics", pa.string()), ("sector", pa.string()),
        ("industry", pa.string()), ("digit", pa.string()), ("basis", pa.string()),
        ("measure", pa.string()), ("units", pa.string()), ("area", pa.string()),
        ("year", pa.int32()), ("value_num", pa.float64()), ("value_str", pa.string()),
        ("extra", pa.string()), ("source", pa.string()),
        ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


def _land_mr_set(dataset: str, uri: str, files: list[str], so: dict) -> tuple[dict, list]:
    """Fetch the named landing files, melt their MachineReadable sheets, write one dataset."""
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    for fn in files:
        raw = _landing_get(f"{PROD_LANDING}{fn}")
        mr = _mr_rows(raw, fn, t0)
        if mr is None:
            print(f"WARN {dataset}: {fn} has no MachineReadable sheet; cell-melting instead", flush=True)
            continue
        rows += mr
    tbl = pa.Table.from_pylist(rows, schema=_mr_schema())
    _write_lance(tbl, uri, so)
    built = _build_indexes(uri, btree=["naics", "year"], bitmap=["source_file", "measure"], so=so)
    cov = {"rows": len(rows), "files": sorted({r['source_file'] for r in rows}),
           "measures": len({r['measure'] for r in rows})}
    return cov, built


def run_mfp(*, so: dict, uris: dict, smoke: bool) -> list[dict]:
    t0 = dt.datetime.now(dt.timezone.utc)
    out = []
    # major sectors
    status, err, cov, built = "error", None, {}, []
    try:
        files = MFP_MAJOR_FILES[:1] if smoke else MFP_MAJOR_FILES
        cov, built = _land_mr_set("bls_mfp_major_sector", uris["bls_mfp_major_sector"], files, so)
        if not smoke:
            assert cov["rows"] >= 3_000, f"mfp major rows {cov['rows']}"
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        out.append(_finish("bls_mfp_major_sector", uris["bls_mfp_major_sector"],
                           f"landing:{PROD_LANDING}", cov.get("rows", 0), built, cov, status, err, t0))
    # detailed industries
    t1 = dt.datetime.now(dt.timezone.utc)
    status, err, cov, built = "error", None, {}, []
    try:
        files = TFP_DETAILED_FILES[:1] if smoke else TFP_DETAILED_FILES
        cov, built = _land_mr_set("bls_tfp_detailed_industries",
                                  uris["bls_tfp_detailed_industries"], files, so)
        if not smoke:
            assert cov["rows"] >= 200_000, f"tfp detailed rows {cov['rows']}"
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        out.append(_finish("bls_tfp_detailed_industries", uris["bls_tfp_detailed_industries"],
                           f"landing:{PROD_LANDING}", cov.get("rows", 0), built, cov, status, err, t1))
    return out


def run_prod_catalog(*, so: dict, uris: dict, smoke: bool) -> list[dict]:
    """Everything else in landing/bls/productivity/: MachineReadable union where present,
    cell-grain otherwise. Per-file fail-open — one bad workbook never blocks the batch."""
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    handled = set(MFP_MAJOR_FILES) | set(TFP_DETAILED_FILES)
    keys = [k for k in _landing_list(PROD_LANDING)
            if k.lower().endswith(".xlsx") and os.path.basename(k) not in handled]
    if smoke:
        keys = keys[:2]
    mr_rows: list[dict] = []
    cell_rows: list[dict] = []
    skipped: list[str] = []
    for k in keys:
        fn = os.path.basename(k)
        try:
            raw = _landing_get(k)
            mr = _mr_rows(raw, fn, t0)
            if mr is not None:
                mr_rows += mr
            else:
                cell_rows += _cell_rows(raw, fn, 0, t0)
        except Exception as exc:  # noqa: BLE001
            skipped.append(fn)
            print(f"WARN prod_catalog {fn}: {exc}", flush=True)
    out = []
    # MR union dataset
    status, err, built = "error", None, []
    try:
        tbl = pa.Table.from_pylist(mr_rows, schema=_mr_schema())
        _write_lance(tbl, uris["bls_productivity_tables"], so)
        built = _build_indexes(uris["bls_productivity_tables"], btree=["naics", "year"],
                               bitmap=["source_file", "measure"], so=so)
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        cov = {"rows": len(mr_rows), "files": len({r['source_file'] for r in mr_rows}),
               "skipped": skipped}
        out.append(_finish("bls_productivity_tables", uris["bls_productivity_tables"],
                           f"landing:{PROD_LANDING}", len(mr_rows), built, cov, status, err, t0))
    # cell-grain fallback dataset
    t1 = dt.datetime.now(dt.timezone.utc)
    status, err, built = "error", None, []
    try:
        tbl = pa.Table.from_pylist(cell_rows, schema=_cells_schema())
        _write_lance(tbl, uris["bls_productivity_cells"], so)
        built = _build_indexes(uris["bls_productivity_cells"], btree=["source_file"],
                               bitmap=["sheet"], so=so)
        status = "success"
    except Exception as exc:  # noqa: BLE001
        err = str(exc)
        raise
    finally:
        cov = {"rows": len(cell_rows), "files": len({r['source_file'] for r in cell_rows})}
        out.append(_finish("bls_productivity_cells", uris["bls_productivity_cells"],
                           f"landing:{PROD_LANDING}", len(cell_rows), built, cov, status, err, t1))
    return out


# ═══════════════════════════════════════════════════════════════════════════════════
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--stream", required=True,
                    choices=["klems", "bea_fa", "aces", "soi", "bds", "qcew", "mfp",
                             "prod_catalog", "all"])
    ap.add_argument("--smoke", action="store_true",
                    help="throwaway URIs under smoke/; first file/year per stream")
    args = ap.parse_args()

    so = _storage_options()
    uris = dict(URIS)
    if args.smoke:
        uris = {k: v.replace("/active/", "/smoke/") for k, v in uris.items()}

    results = []
    run = args.stream
    if run in ("klems", "all"):
        results.append(run_klems(so=so, uri=uris["bea_bls_klems"], smoke=args.smoke))
    if run in ("bea_fa", "all"):
        results.append(run_bea_fa(so=so, uri=uris["bea_fixed_assets_detail"], smoke=args.smoke))
    if run in ("aces", "all"):
        results.append(run_aces(so=so, uri=uris["census_aces_capex"], smoke=args.smoke))
    if run in ("soi", "all"):
        results.append(run_soi(so=so, uri=uris["irs_soi_corp_industry"], smoke=args.smoke))
    if run in ("bds", "all"):
        results.append(run_bds(so=so, uri=uris["census_bds_firm_dynamics"], smoke=args.smoke))
    if run in ("mfp", "all"):
        results += run_mfp(so=so, uris=uris, smoke=args.smoke)
    if run in ("prod_catalog", "all"):
        results += run_prod_catalog(so=so, uris=uris, smoke=args.smoke)
    if run in ("qcew", "all"):
        results.append(run_qcew(so=so, uri=uris["bls_qcew_annual"], smoke=args.smoke))

    print("\n=== BATCH RESULT ===", flush=True)
    for r in results:
        print(f"  {r['dataset']}: rows={r['rows']} status={r['status']}", flush=True)
    if any(r["status"] != "success" for r in results):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
