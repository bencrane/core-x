"""Labor-Share-of-Revenue Stack — Census SUSB + BEA Value-Added + BLS ECEC → Gen-3 Lance SoR.

Supplies the ONE missing scalar in the NAICS×PSC award→labor decomposition: the fraction of
an award dollar that is labor at all. The within-labor MIX (which categories, at what wage) is
already live (naics_psc_labor_dim / naics_psc_labor_profile_categories); this module lands the
SHARE and its calibration so downstream can close:

    expected labor $ by category = award_$ × labor_share × category_mix

Three keyless public statistical streams, deterministic parses, Pattern A (direct hydration:
ephemeral download → PyArrow project → lance.write_dataset → s3://data-sink/active/<name>/).
No RisingWave, no Postgres source tables, no catalog layer. Raw stays lossless (verbatim
strings + suppression/noise flags preserved; derived ratios are ADDITIONAL columns).

  STREAM 1 — Census SUSB 2022 (--stream susb)  →  active/census_susb_naics_payroll_receipts/
    Static xlsx (keyless, browser-UA; the Census API is now key-gated — do NOT use it).
    us_6digitnaics_rcptsize_2022.xlsx, sheet 'US 6-digit NAICS', header on row 3, data row 4+.
    Grain: NAICS × enterprise-size class. Derived: payroll_share = annual_payroll / receipts.
    Payroll & Receipts are in $1,000s (Economic-Census-year only; 2022 hardcoded as ref_year).

  STREAM 2 — BEA GDP-by-Industry value added (--stream bea)  →  active/bea_industry_value_added/
    Static zip (keyless). Two annual dollar tables, long-form:
      ValueAdded.xlsx::TVA113-A  "Components of Value Added by Industry" [$M, 1997–2024]
        per industry: value_added header + 3 sub-rows (compensation_of_employees /
        taxes_on_production_less_subsidies / gross_operating_surplus)
      GrossOutput.xlsx::TGO105-A "Gross Output by Industry"            [$M, 1997–2025]
        the denominator (component='gross_output')
    Derived rows (component='derived_comp_share', latest 3 common years):
        comp_share_of_output = CoE(industry, year) / GrossOutput(industry, year)
    The GDPbyIndustry summary workbooks expose Line + Name (no alphanumeric industry code),
    so bea_industry_line is the stable join key; bea_industry_code lands NULL. The cross-table
    join is by trimmed industry_name with a single economy-wide alias (the value-added table's
    top line is "Gross domestic product", gross output's is "All industries"). NAICS↔BEA
    concordance is deliberately OUT OF SCOPE (downstream composition work).

  STREAM 3 — BLS ECEC (--stream ecec)  →  active/bls_ecec_costs/  +  active/bls_ecec_burden/
    BLS Public Data API v2 POST (keyless; 25 series/query, 500/day — the download.bls.gov flat
    files 403 even with a browser UA, so the API is the only path). ECEC series ID format
    (authoritative, bls.gov/ecec/factsheets/ecec-series-id-guide.htm):
        CM U <owner:1> <estimate:2> <industry:4> <occupation:3> <subcell:3> <datatype:1>
    Landed: private (owner=2), all-occ (000), all-worker (000), cost-per-hour (D), estimate
    01 (total compensation) and 02 (wages & salaries), across the published major industry
    groups. Every landed series returns REQUEST_SUCCEEDED with ≥4 quarters (others are dropped
    and logged). Derived: burden_multiplier = total_comp / wages_salaries per (ownership,
    industry_group) at the latest common quarter — the multiplier that converts a payroll share
    into a fully-loaded labor-cost share.

Ledger: ops.labor_share_runs (psycopg via HQX_DB_URL_POOLED; create-if-not-exists inline; an
audit-write failure warns, never masks a good load).

    doppler run -p core-x -c prd -- uv run --with pylance --with pyarrow --with duckdb \\
      --with requests --with boto3 --with openpyxl --with 'psycopg[binary]' \\
      python -m pipelines.reference.labor_share_ingest --stream all --smoke   # throwaway URIs
    ... python -m pipelines.reference.labor_share_ingest --stream all                 # full
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import os
import time
import zipfile

# Reuse the fleet R2/index plumbing verbatim (do not reimplement).
from pipelines.bls.ingest import (  # noqa: E402
    DATA_STORAGE_VERSION,
    MAX_BYTES_PER_FILE,
    MAX_ROWS_PER_FILE,
    _build_indexes,
    _s3_client,
    _storage_options,
)

BUCKET = "data-sink"
SOURCE_TAG = "labor_share_ingest"

# ── dataset URIs (overridable via env; --smoke rewrites to _smoke_ prefixes) ─────────
SUSB_URI = os.environ.get("SUSB_LANCE_URI", f"s3://{BUCKET}/active/census_susb_naics_payroll_receipts/")
BEA_URI = os.environ.get("BEA_LANCE_URI", f"s3://{BUCKET}/active/bea_industry_value_added/")
ECEC_COSTS_URI = os.environ.get("ECEC_COSTS_LANCE_URI", f"s3://{BUCKET}/active/bls_ecec_costs/")
ECEC_BURDEN_URI = os.environ.get("ECEC_BURDEN_LANCE_URI", f"s3://{BUCKET}/active/bls_ecec_burden/")

SUSB_URL = "https://www2.census.gov/programs-surveys/susb/tables/2022/us_6digitnaics_rcptsize_2022.xlsx"
BEA_ZIP_URL = "https://apps.bea.gov/industry/Release/ZIP/GdpByInd.zip"
BLS_API_URL = "https://api.bls.gov/publicAPI/v2/timeseries/data/"

SUSB_REF_YEAR = 2022

_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/126 Safari/537.36")


# ── ECEC code tables (authoritative, from the BLS ECEC series-id guide) ──────────────
ECEC_OWNERSHIP = {"1": "civilian workers", "2": "private industry workers",
                  "3": "state and local government workers"}
ECEC_ESTIMATE = {"01": "total compensation", "02": "wages and salaries"}
# Published major industry groups (EEEE → name). Detailed 4-digit-NAICS subcells are excluded.
ECEC_INDUSTRY = {
    "0000": "all industries",
    "G000": "goods-producing industries",
    "S000": "service-providing industries",
    "2300": "construction",
    "3000": "manufacturing",
    "4400": "utilities",
    "4000": "trade, transportation, and utilities",
    "4110": "wholesale trade",
    "4120": "retail trade",
    "4300": "transportation and warehousing",
    "5100": "information",
    "5200": "finance and insurance",
    "520A": "financial activities",
    "5300": "real estate and rental and leasing",
    "5400": "professional, scientific, and technical services",
    "540A": "professional and business services",
    "5600": "administrative and support and waste management and remediation services",
    "6000": "education and health services",
    "6100": "educational services",
    "6200": "health care and social assistance",
    "6220": "hospitals",
    "7000": "leisure and hospitality",
    "7200": "accommodation and food services",
    "8100": "other services (except public administration)",
}
ECEC_OWNER = "2"          # private industry
ECEC_OCC = "000"          # all occupations
ECEC_SUBCELL = "000"      # all workers
ECEC_DATATYPE = "D"       # cost per hour worked, current dollars
ECEC_START_YEAR = "2020"  # land 2020Q1 forward


def _ecec_series_id(estimate: str, industry: str) -> str:
    return f"CM U{ECEC_OWNER}{estimate}{industry}{ECEC_OCC}{ECEC_SUBCELL}{ECEC_DATATYPE}".replace(" ", "")


# ── HTTP ────────────────────────────────────────────────────────────────────────────
def _download(url: str, *, timeout: int = 240) -> bytes:
    """GET raw bytes with a browser UA; retry transient failures with backoff."""
    import requests

    last = None
    for attempt in range(6):
        try:
            resp = requests.get(url, headers={"User-Agent": _UA}, timeout=timeout)
        except requests.RequestException as exc:  # noqa: PERF203
            last = exc
            time.sleep(min(30, 2 ** attempt))
            continue
        if resp.status_code == 200:
            return resp.content
        if resp.status_code in (429, 500, 502, 503, 504):
            last = RuntimeError(f"HTTP {resp.status_code}")
            time.sleep(min(60, 2 ** attempt))
            continue
        raise RuntimeError(f"GET {url} HTTP {resp.status_code}: {(resp.text or '')[:200]}")
    raise RuntimeError(f"GET {url}: retries exhausted ({last}).")


# ── coercion helpers ────────────────────────────────────────────────────────────────
def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _naics_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip() or None


def _int(v) -> int | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int,)):
        return int(v)
    if isinstance(v, float):
        return int(v) if v.is_integer() else int(round(v))
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s or s in {"(D)", "(NA)", "...", "…", "*", "**", "-", "--"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _naics_level(naics: str | None) -> int | None:
    if naics is None:
        return None
    if naics == "--":
        return 0
    if "-" in naics:  # NAICS sector ranges (31-33, 44-45, 48-49) are 2-digit sector level
        return 2
    return len(naics)


# ── ops.labor_share_runs ledger ─────────────────────────────────────────────────────
def _record_run(dataset: str, uri: str, source_url: str, rows: int, built: list,
                coverage: dict, status: str, error: str | None,
                started_at, completed_at) -> None:
    dsn = os.environ.get("HQX_DB_URL_POOLED")
    if not dsn:
        print("WARN: HQX_DB_URL_POOLED not set; skipping ops.labor_share_runs write.", flush=True)
        return
    try:
        import psycopg

        with psycopg.connect(dsn) as conn, conn.cursor() as cur:
            cur.execute("CREATE SCHEMA IF NOT EXISTS ops;")
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS ops.labor_share_runs (
                    id bigint GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                    dataset text NOT NULL, dataset_uri text, source_url text,
                    rows_processed bigint, indexes_built text[], coverage jsonb,
                    status text NOT NULL, error text,
                    started_at timestamptz, completed_at timestamptz,
                    recorded_at timestamptz NOT NULL DEFAULT now());
                """
            )
            cur.execute(
                """
                INSERT INTO ops.labor_share_runs
                    (dataset, dataset_uri, source_url, rows_processed, indexes_built,
                     coverage, status, error, started_at, completed_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (dataset, uri, source_url, rows, built, json.dumps(coverage), status, error,
                 started_at, completed_at),
            )
            conn.commit()
    except Exception as exc:  # noqa: BLE001 — audit must never mask a good load
        print(f"WARN: ops.labor_share_runs write failed: {exc}", flush=True)


def _write_lance(table, uri: str, storage_options: dict) -> None:
    import lance

    lance.write_dataset(table, uri, mode="overwrite",
                        data_storage_version=DATA_STORAGE_VERSION,
                        max_rows_per_file=MAX_ROWS_PER_FILE, max_bytes_per_file=MAX_BYTES_PER_FILE,
                        storage_options=storage_options)


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM 1 — Census SUSB 2022
# ═══════════════════════════════════════════════════════════════════════════════════
def _susb_schema():
    import pyarrow as pa
    return pa.schema([
        ("naics", pa.string()), ("naics_level", pa.int32()), ("naics_description", pa.string()),
        ("size_class", pa.string()), ("firms", pa.int64()), ("establishments", pa.int64()),
        ("employment", pa.int64()), ("employment_flag", pa.string()),
        ("annual_payroll_k", pa.int64()), ("payroll_flag", pa.string()),
        ("receipts_k", pa.int64()), ("receipts_flag", pa.string()),
        ("payroll_share", pa.float64()), ("ref_year", pa.int32()),
        ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


def run_susb(*, storage_options: dict, uri: str = SUSB_URI) -> dict:
    import openpyxl
    import pyarrow as pa

    started_at = dt.datetime.now(dt.timezone.utc)
    ingested_at = started_at
    status, error_text, built = "error", None, []
    rows: list[dict] = []
    try:
        raw = _download(SUSB_URL)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb["US 6-digit NAICS"]
        for r in ws.iter_rows(min_row=4, values_only=True):  # rows 1-2 banners, row 3 header
            if r[0] is None and r[2] is None:
                continue
            naics = _naics_str(r[0])
            payroll_k = _int(r[7])
            receipts_k = _int(r[9])
            share = (payroll_k / receipts_k) if (payroll_k is not None and receipts_k) else None
            rows.append({
                "naics": naics, "naics_level": _naics_level(naics),
                "naics_description": _s(r[1]), "size_class": _s(r[2]),
                "firms": _int(r[3]), "establishments": _int(r[4]), "employment": _int(r[5]),
                "employment_flag": _s(r[6]), "annual_payroll_k": payroll_k,
                "payroll_flag": _s(r[8]), "receipts_k": receipts_k, "receipts_flag": _s(r[10]),
                "payroll_share": share, "ref_year": SUSB_REF_YEAR,
                "source": f"{SOURCE_TAG}:census_susb_2022", "ingested_at": ingested_at,
            })
        wb.close()

        tbl = pa.Table.from_pylist(rows, schema=_susb_schema())
        _write_lance(tbl, uri, storage_options)
        built = _build_indexes(uri, btree=["naics"],
                               bitmap=["size_class", "naics_level", "ref_year"], so=storage_options)
        status = "success"
    except Exception as exc:  # noqa: BLE001
        error_text = str(exc)
        print(f"FATAL susb: {exc}", flush=True)
        raise
    finally:
        completed_at = dt.datetime.now(dt.timezone.utc)
        n6 = len({r["naics"] for r in rows if r["naics_level"] == 6})
        econ = next((r for r in rows if r["naics"] == "--" and r["size_class"] == "01: Total"), None)
        cov = {
            "rows": len(rows), "distinct_naics_6digit": n6,
            "distinct_size_class": len({r["size_class"] for r in rows}),
            "economy_payroll_k": econ["annual_payroll_k"] if econ else None,
            "economy_receipts_k": econ["receipts_k"] if econ else None,
            "economy_payroll_share": round(econ["payroll_share"], 6) if econ and econ["payroll_share"] else None,
        }
        _record_run("census_susb_naics_payroll_receipts", uri, SUSB_URL, len(rows), built,
                    cov, status, error_text, started_at, completed_at)
        print(f"SUSB SUMMARY: {cov} status={status}", flush=True)
    return {"dataset": "census_susb_naics_payroll_receipts", "rows": len(rows),
            "indexes": built, "coverage": cov, "status": status}


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM 2 — BEA GDP-by-Industry value added
# ═══════════════════════════════════════════════════════════════════════════════════
_BEA_COMPONENTS = {
    "compensation of employees": "compensation_of_employees",
    "taxes on production and imports less subsidies": "taxes_on_production_less_subsidies",
    "gross operating surplus": "gross_operating_surplus",
}
# economy-wide alias: value-added table top line vs gross-output table top line
_BEA_ECON_ALIAS = {"gross domestic product": "all industries"}


def _bea_schema():
    import pyarrow as pa
    return pa.schema([
        ("table", pa.string()), ("bea_industry_line", pa.int32()),
        ("bea_industry_code", pa.string()), ("industry_name", pa.string()),
        ("component", pa.string()), ("year", pa.int32()), ("value", pa.float64()),
        ("comp_share_of_output", pa.float64()),
        ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


def _bea_year_cols(ws) -> list[tuple[int, int]]:
    """Header on row 8 (col D+): return [(col_index, year), ...] for 4-digit year labels."""
    header = next(ws.iter_rows(min_row=8, max_row=8, values_only=True))
    out = []
    for j, cell in enumerate(header):
        s = _s(cell)
        if s and len(s) == 4 and s.isdigit():
            out.append((j, int(s)))
    return out


def _bea_parse_components(ws, table_code: str, ingested_at):
    """TVA113-A: industry header row = value_added; 3 sub-rows = CoE/Taxes/GOS.
    Returns (long_rows, coe_by_key_year, va_meta_by_key). Component rows inherit the
    most-recent industry header's line + name."""
    ycols = _bea_year_cols(ws)
    rows, coe, va_meta = [], {}, {}
    cur_line, cur_name, cur_key = None, None, None
    for r in ws.iter_rows(min_row=9, values_only=True):
        line = _int(r[0])
        name = _s(r[1])
        if name is None:
            continue
        comp = _BEA_COMPONENTS.get(name.lower())
        if comp is None:  # industry header row → value_added
            cur_line, cur_name = line, name
            cur_key = _BEA_ECON_ALIAS.get(name.lower(), name.lower())
            component = "value_added"
            va_meta[cur_key] = {"line": cur_line, "name": cur_name}
        else:
            component = comp
        if cur_line is None:
            continue
        for j, year in ycols:
            val = _float(r[j]) if j < len(r) else None
            if val is None:
                continue
            rows.append({
                "table": table_code, "bea_industry_line": cur_line, "bea_industry_code": None,
                "industry_name": cur_name, "component": component, "year": year, "value": val,
                "comp_share_of_output": None, "source": f"{SOURCE_TAG}:bea_gdpbyind",
                "ingested_at": ingested_at,
            })
            if component == "compensation_of_employees":
                coe[(cur_key, year)] = val
    return rows, coe, va_meta


def _bea_parse_grossoutput(ws, table_code: str, ingested_at):
    """TGO105-A: one row per industry (component='gross_output'). Returns (rows, go_by_key_year)."""
    ycols = _bea_year_cols(ws)
    rows, go = [], {}
    for r in ws.iter_rows(min_row=9, values_only=True):
        line = _int(r[0])
        name = _s(r[1])
        if name is None:
            continue
        key = _BEA_ECON_ALIAS.get(name.lower(), name.lower())
        for j, year in ycols:
            val = _float(r[j]) if j < len(r) else None
            if val is None:
                continue
            rows.append({
                "table": table_code, "bea_industry_line": line, "bea_industry_code": None,
                "industry_name": name, "component": "gross_output", "year": year, "value": val,
                "comp_share_of_output": None, "source": f"{SOURCE_TAG}:bea_gdpbyind",
                "ingested_at": ingested_at,
            })
            go[(key, year)] = val
    return rows, go


def run_bea(*, storage_options: dict, uri: str = BEA_URI, derived_years: int = 3) -> dict:
    import openpyxl
    import pyarrow as pa

    started_at = dt.datetime.now(dt.timezone.utc)
    ingested_at = started_at
    status, error_text, built = "error", None, []
    rows: list[dict] = []
    cov: dict = {}
    try:
        zf = zipfile.ZipFile(io.BytesIO(_download(BEA_ZIP_URL)))
        va_wb = openpyxl.load_workbook(io.BytesIO(zf.read("ValueAdded.xlsx")), read_only=True, data_only=True)
        go_wb = openpyxl.load_workbook(io.BytesIO(zf.read("GrossOutput.xlsx")), read_only=True, data_only=True)
        va_rows, coe, va_meta = _bea_parse_components(va_wb["TVA113-A"], "TVA113", ingested_at)
        go_rows, go = _bea_parse_grossoutput(go_wb["TGO105-A"], "TGO105", ingested_at)
        va_wb.close(); go_wb.close()
        rows = va_rows + go_rows

        # Derived comp_share_of_output = CoE / GrossOutput, joined by industry key, latest N common years.
        keys = {k for (k, _y) in coe} & {k for (k, _y) in go}
        derived, per_share = [], []
        for key in keys:
            common = sorted({y for (kk, y) in coe if kk == key} & {y for (kk, y) in go if kk == key})
            for year in common[-derived_years:]:
                den = go[(key, year)]
                if not den:
                    continue
                share = coe[(key, year)] / den
                meta = va_meta.get(key, {})
                derived.append({
                    "table": "DERIVED", "bea_industry_line": meta.get("line"),
                    "bea_industry_code": None, "industry_name": meta.get("name", key),
                    "component": "derived_comp_share", "year": year, "value": share,
                    "comp_share_of_output": share, "source": f"{SOURCE_TAG}:bea_gdpbyind",
                    "ingested_at": ingested_at,
                })
                per_share.append(share)
        rows += derived

        tbl = pa.Table.from_pylist(rows, schema=_bea_schema())
        _write_lance(tbl, uri, storage_options)
        built = _build_indexes(uri, btree=["bea_industry_line", "industry_name"],
                               bitmap=["component", "table", "year"], so=storage_options)
        status = "success"
    except Exception as exc:  # noqa: BLE001
        error_text = str(exc)
        print(f"FATAL bea: {exc}", flush=True)
        raise
    finally:
        completed_at = dt.datetime.now(dt.timezone.utc)
        raw_years = [r["year"] for r in rows if r["table"] != "DERIVED"]
        latest = max(raw_years) if raw_years else None
        go_latest = [r for r in rows if r["table"] == "TGO105" and r["year"] == latest]
        econ = next((r for r in rows if r["table"] == "DERIVED"
                     and r["industry_name"] and r["industry_name"].strip().lower() == "gross domestic product"
                     and r["year"] == max((d["year"] for d in rows if d["table"] == "DERIVED"), default=0)), None)
        shares = [r["comp_share_of_output"] for r in rows if r["table"] == "DERIVED"]
        cov = {
            "rows": len(rows), "raw_rows": len(rows) - sum(1 for r in rows if r["table"] == "DERIVED"),
            "derived_rows": sum(1 for r in rows if r["table"] == "DERIVED"),
            "components": sorted({r["component"] for r in rows}),
            "latest_year": latest,
            "industry_lines_latest_go": len({r["industry_name"] for r in go_latest}),
            "economy_comp_share_latest": round(econ["comp_share_of_output"], 6) if econ else None,
            "share_min": round(min(shares), 6) if shares else None,
            "share_max": round(max(shares), 6) if shares else None,
        }
        _record_run("bea_industry_value_added", uri, BEA_ZIP_URL, len(rows), built,
                    cov, status, error_text, started_at, completed_at)
        print(f"BEA SUMMARY: {cov} status={status}", flush=True)
    return {"dataset": "bea_industry_value_added", "rows": len(rows),
            "indexes": built, "coverage": cov, "status": status}


# ═══════════════════════════════════════════════════════════════════════════════════
# STREAM 3 — BLS ECEC (costs + burden)
# ═══════════════════════════════════════════════════════════════════════════════════
def _bls_post(series_ids: list[str], start_year: str, end_year: str) -> dict:
    import requests

    body = json.dumps({"seriesid": series_ids, "startyear": start_year, "endyear": end_year})
    hdrs = {"Content-Type": "application/json", "User-Agent": _UA}
    for attempt in range(6):
        try:
            resp = requests.post(BLS_API_URL, data=body, headers=hdrs, timeout=180)
        except requests.RequestException:
            time.sleep(min(30, 2 ** attempt)); continue
        if resp.status_code == 200:
            try:
                return resp.json()
            except ValueError:
                time.sleep(min(60, 2 ** attempt)); continue
        if resp.status_code in (429, 500, 502, 503, 504):
            time.sleep(min(60, 5 * 2 ** attempt)); continue
        raise RuntimeError(f"BLS POST HTTP {resp.status_code}: {(resp.text or '')[:200]}")
    raise RuntimeError("BLS POST: retries exhausted.")


def _ecec_costs_schema():
    import pyarrow as pa
    return pa.schema([
        ("series_id", pa.string()), ("series_title", pa.string()), ("ownership", pa.string()),
        ("component", pa.string()), ("industry_group", pa.string()), ("occupation_group", pa.string()),
        ("year", pa.int32()), ("period", pa.string()), ("value", pa.float64()),
        ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


def _ecec_burden_schema():
    import pyarrow as pa
    return pa.schema([
        ("ownership", pa.string()), ("industry_group", pa.string()), ("quarter", pa.string()),
        ("total_comp", pa.float64()), ("wages_salaries", pa.float64()),
        ("burden_multiplier", pa.float64()),
        ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


def run_ecec(*, storage_options: dict, costs_uri: str = ECEC_COSTS_URI,
             burden_uri: str = ECEC_BURDEN_URI, end_year: str | None = None) -> dict:
    import pyarrow as pa

    started_at = dt.datetime.now(dt.timezone.utc)
    ingested_at = started_at
    end_year = end_year or str(started_at.year)

    # Build the candidate series catalog: private × {total comp, W&S} × major industry groups.
    catalog: dict[str, dict] = {}
    for estimate in ("01", "02"):
        for eeee, ind_name in ECEC_INDUSTRY.items():
            sid = _ecec_series_id(estimate, eeee)
            catalog[sid] = {
                "series_id": sid, "ownership": ECEC_OWNERSHIP[ECEC_OWNER],
                "component": ECEC_ESTIMATE[estimate], "estimate": estimate,
                "industry_group": ind_name, "industry_code": eeee,
                "occupation_group": "all workers",
                "series_title": (f"ECEC: {ECEC_ESTIMATE[estimate]}, {ECEC_OWNERSHIP[ECEC_OWNER]}, "
                                 f"{ind_name}, all occupations, all workers, cost per hour worked"),
            }

    status, error_text, built = "error", None, []
    cost_rows: list[dict] = []
    burden_rows: list[dict] = []
    landed_series: set[str] = set()
    dropped: list[str] = []
    cov: dict = {}
    try:
        ids = list(catalog)
        for i in range(0, len(ids), 25):  # keyless API caps 25 series/query
            chunk = ids[i:i + 25]
            resp = _bls_post(chunk, ECEC_START_YEAR, end_year)
            if resp.get("status") != "REQUEST_SUCCEEDED":
                raise RuntimeError(f"BLS batch status={resp.get('status')} msgs={resp.get('message')}")
            for s in resp.get("Results", {}).get("series", []):
                sid = s["seriesID"]
                meta = catalog[sid]
                data = [d for d in s.get("data", []) if d.get("period", "").startswith("Q")]
                if len(data) < 4:  # require ≥4 quarters
                    dropped.append(sid)
                    continue
                landed_series.add(sid)
                for d in data:
                    val = _float(d.get("value"))
                    if val is None:
                        continue
                    cost_rows.append({
                        "series_id": sid, "series_title": meta["series_title"],
                        "ownership": meta["ownership"], "component": meta["component"],
                        "industry_group": meta["industry_group"],
                        "occupation_group": meta["occupation_group"],
                        "year": _int(d["year"]), "period": d["period"], "value": val,
                        "source": f"{SOURCE_TAG}:bls_ecec_v2", "ingested_at": ingested_at,
                    })
            time.sleep(0.5)

        # Burden = total_comp / wages_salaries per industry group at the latest common quarter.
        def _latest(component: str, eeee: str):
            pts = [c for c in cost_rows if c["component"] == component
                   and c["industry_group"] == ECEC_INDUSTRY[eeee]]
            if not pts:
                return None
            best = max(pts, key=lambda c: (c["year"], c["period"]))
            return best

        for eeee, ind_name in ECEC_INDUSTRY.items():
            tc = _latest("total compensation", eeee)
            ws = _latest("wages and salaries", eeee)
            if not tc or not ws or not ws["value"]:
                continue
            # align to a common quarter if the latest differ
            if (tc["year"], tc["period"]) != (ws["year"], ws["period"]):
                tc_by = {(c["year"], c["period"]): c for c in cost_rows
                         if c["component"] == "total compensation" and c["industry_group"] == ind_name}
                ws_by = {(c["year"], c["period"]): c for c in cost_rows
                         if c["component"] == "wages and salaries" and c["industry_group"] == ind_name}
                common = sorted(set(tc_by) & set(ws_by))
                if not common:
                    continue
                q = common[-1]
                tc, ws = tc_by[q], ws_by[q]
            burden_rows.append({
                "ownership": ECEC_OWNERSHIP[ECEC_OWNER], "industry_group": ind_name,
                "quarter": f"{tc['year']} {tc['period']}", "total_comp": tc["value"],
                "wages_salaries": ws["value"], "burden_multiplier": tc["value"] / ws["value"],
                "source": f"{SOURCE_TAG}:bls_ecec_v2", "ingested_at": ingested_at,
            })

        if not cost_rows:
            raise RuntimeError("ECEC: no series landed (all candidates dropped).")
        _write_lance(pa.Table.from_pylist(cost_rows, schema=_ecec_costs_schema()), costs_uri, storage_options)
        built = _build_indexes(costs_uri, btree=["series_id"],
                               bitmap=["industry_group", "component", "ownership", "period"],
                               so=storage_options)
        _write_lance(pa.Table.from_pylist(burden_rows, schema=_ecec_burden_schema()), burden_uri, storage_options)
        built += ["burden:" + b for b in _build_indexes(
            burden_uri, btree=["industry_group"], bitmap=["ownership"], so=storage_options)]
        status = "success"
    except Exception as exc:  # noqa: BLE001
        error_text = str(exc)
        print(f"FATAL ecec: {exc}", flush=True)
        raise
    finally:
        completed_at = dt.datetime.now(dt.timezone.utc)
        anchor = next((c for c in cost_rows if c["series_id"] == "CMU2010000000000D"), None)
        anchor_latest = (max([c for c in cost_rows if c["series_id"] == "CMU2010000000000D"],
                             key=lambda c: (c["year"], c["period"])) if anchor else None)
        econ_burden = next((b for b in burden_rows if b["industry_group"] == "all industries"), None)
        mults = [b["burden_multiplier"] for b in burden_rows]
        cov = {
            "candidate_series": len(catalog), "landed_series": len(landed_series),
            "dropped_series": dropped, "cost_rows": len(cost_rows), "burden_rows": len(burden_rows),
            "anchor_latest": (f"{anchor_latest['year']} {anchor_latest['period']} = {anchor_latest['value']}"
                              if anchor_latest else None),
            "economy_private_multiplier": round(econ_burden["burden_multiplier"], 4) if econ_burden else None,
            "burden_min": round(min(mults), 4) if mults else None,
            "burden_max": round(max(mults), 4) if mults else None,
        }
        _record_run("bls_ecec_costs", costs_uri, BLS_API_URL, len(cost_rows), built,
                    cov, status, error_text, started_at, completed_at)
        _record_run("bls_ecec_burden", burden_uri, BLS_API_URL, len(burden_rows), built,
                    cov, status, error_text, started_at, completed_at)
        print(f"ECEC SUMMARY: {cov} status={status}", flush=True)
    return {"dataset": "bls_ecec_costs+burden", "cost_rows": len(cost_rows),
            "burden_rows": len(burden_rows), "indexes": built, "coverage": cov, "status": status}


# ═══════════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════════
def _smoke_uri(uri: str) -> str:
    base = uri.rstrip("/").rsplit("/", 1)[-1]
    return f"s3://{BUCKET}/active/_smoke_{base}/"


def _delete_prefix(uri: str) -> None:
    s3 = _s3_client()
    prefix = uri.split(f"{BUCKET}/", 1)[1]
    paginator = s3.get_paginator("list_objects_v2")
    keys = [o["Key"] for page in paginator.paginate(Bucket=BUCKET, Prefix=prefix)
            for o in page.get("Contents", [])]
    for i in range(0, len(keys), 1000):
        s3.delete_objects(Bucket=BUCKET, Delete={"Objects": [{"Key": k} for k in keys[i:i + 1000]]})
    print(f"  smoke cleanup: deleted {len(keys)} objects under {prefix}", flush=True)


def _cli() -> None:
    p = argparse.ArgumentParser(description="Labor-share-of-revenue stack ingest (SUSB + BEA + ECEC → Lance).")
    p.add_argument("--stream", choices=["susb", "bea", "ecec", "all"], required=True)
    p.add_argument("--smoke", action="store_true", help="write to throwaway _smoke_ URIs and delete them after.")
    a = p.parse_args()

    so = _storage_options()
    streams = ["susb", "bea", "ecec"] if a.stream == "all" else [a.stream]
    uris = {"susb": SUSB_URI, "bea": BEA_URI, "ecec_costs": ECEC_COSTS_URI, "ecec_burden": ECEC_BURDEN_URI}
    if a.smoke:
        uris = {k: _smoke_uri(v) for k, v in uris.items()}

    results = []
    for s in streams:
        if s == "susb":
            results.append(run_susb(storage_options=so, uri=uris["susb"]))
        elif s == "bea":
            results.append(run_bea(storage_options=so, uri=uris["bea"]))
        elif s == "ecec":
            results.append(run_ecec(storage_options=so, costs_uri=uris["ecec_costs"],
                                    burden_uri=uris["ecec_burden"]))

    if a.smoke:
        for s in streams:
            if s == "susb":
                _delete_prefix(uris["susb"])
            elif s == "bea":
                _delete_prefix(uris["bea"])
            elif s == "ecec":
                _delete_prefix(uris["ecec_costs"]); _delete_prefix(uris["ecec_burden"])

    print("\n=== labor-share ingest summary ===", flush=True)
    for r in results:
        print(f"  {r}", flush=True)


if __name__ == "__main__":
    _cli()
