"""Federal appropriations & budget authority ingest — OMB Public Budget Database +
USAspending Budgetary Resources → Lance system of record (Gen-3).

This lands the *appropriated* side of the federal dollar, which the plane otherwise lacks
entirely: every existing number is an obligation (money committed on a contract); nothing
measured budget authority (money Congress made available). Ten Lance datasets under
``s3://data-sink/active/`` at account grain and agency grain, plus one derived
appropriated-vs-obligated FY reconciliation.

    doppler run -p core-x -c prd -- uv run --with pylance --with pyarrow --with duckdb \\
      --with openpyxl --with requests --with boto3 --with 'psycopg[binary]' \\
      python -m pipelines.reference.federal_appropriations_ingest --stream all --smoke   # smoke
    ... --stream all                                                                     # full

STREAM → DATASET
    omb_budauth        → active/omb_budget_authority/            account × year  (OMB xlsx)
    omb_outlays        → active/omb_outlays/                     account × year  (OMB xlsx)
    usa_account_balances → active/usaspending_federal_account_balances/  (File A, federal grain)
                        +  active/usaspending_treasury_account_balances/ (File A, TAS grain)
    usa_agency         → active/usaspending_agency_fy/           agency × FY
    usa_agency_periods → active/usaspending_agency_period_obligations/  agency × FY × period
    usa_total_resources→ active/usaspending_total_budgetary_resources/  FY × period (gov-wide)
    usa_account_roster → active/usaspending_federal_account_roster/     federal account (crosswalk)
    usa_defc           → active/usaspending_def_codes/           DEFC registry
    derived            → active/approp_vs_obligated_fy/           FY reconciliation (reads #1,#2,#3)

RATE DISCIPLINE — every network call routes through pipelines/_lib/rate_governor.py
    (host-scoped token bucket + warm-up ramp + circuit breaker + path checkpoint). These
    publishers return no rate-limit headers and go straight to an IP-scoped block; the
    governor makes over-fetching structurally impossible. One governor per host, driven
    single-threaded. whitehouse.gov gets a browser UA (it 403s a bare client); usaspending
    gets the honest descriptive UA. See the governor docstring for the binding contract.

Fidelity: agency/bureau/account/subfunction codes land verbatim as strings (leading zeros
are load-bearing). Negative values are real (receipt / offsetting accounts) — never abs().
Estimates are flagged, never mixed with actuals: is_estimate is derived from the workbook
vintage in the FILENAME (budauth_fy{YYYY} → budget_year=YYYY; is_estimate = year>=YYYY-1),
never hardcoded. The 1976 transition quarter lands with a non-colliding sentinel
(year=-1976, is_transition_quarter=true) so no rollup conflates a 3-month stub into FY1976.

Ledger: ops.federal_appropriations_ingest_runs (HQX_DB_URL_POOLED; create-if-not-exists
inline; status CHECK is the canonical 3 values; throttle vocabulary rides free-text
disposition). Sources register in ops.data_source_catalog (ON CONFLICT DO NOTHING).
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import os
import re
import tempfile
import time
import uuid
import zipfile

# Cyberduck/R2 multipart checksums — validate only when explicitly requested (fleet const).
os.environ.setdefault("AWS_RESPONSE_CHECKSUM_VALIDATION", "when_required")
os.environ.setdefault("AWS_REQUEST_CHECKSUM_CALCULATION", "when_required")

# Reuse fleet R2/index plumbing verbatim (do NOT reimplement).
from pipelines.bls.ingest import (  # noqa: E402
    DATA_STORAGE_VERSION,
    MAX_BYTES_PER_FILE,
    MAX_ROWS_PER_FILE,
    _build_indexes,
    _storage_options,
)
from pipelines._lib.rate_governor import RateGovernor, ThrottleHalt  # noqa: E402

BUCKET = "data-sink"
SOURCE_TAG = "federal_appropriations_ingest"

URIS = {
    "omb_budget_authority": f"s3://{BUCKET}/active/omb_budget_authority/",
    "omb_outlays": f"s3://{BUCKET}/active/omb_outlays/",
    "usaspending_federal_account_balances": f"s3://{BUCKET}/active/usaspending_federal_account_balances/",
    "usaspending_treasury_account_balances": f"s3://{BUCKET}/active/usaspending_treasury_account_balances/",
    "usaspending_agency_fy": f"s3://{BUCKET}/active/usaspending_agency_fy/",
    "usaspending_agency_period_obligations": f"s3://{BUCKET}/active/usaspending_agency_period_obligations/",
    "usaspending_total_budgetary_resources": f"s3://{BUCKET}/active/usaspending_total_budgetary_resources/",
    "usaspending_federal_account_roster": f"s3://{BUCKET}/active/usaspending_federal_account_roster/",
    "usaspending_def_codes": f"s3://{BUCKET}/active/usaspending_def_codes/",
    "approp_vs_obligated_fy": f"s3://{BUCKET}/active/approp_vs_obligated_fy/",
}

# ── upstream endpoints ───────────────────────────────────────────────────────────────
OMB_SUPPLEMENTAL = ("https://www.whitehouse.gov/omb/information-resources/budget/"
                    "supplemental-materials/")
# verified-live FY2027 release (2026-07-27) — used only as a fallback if discovery misses.
OMB_FALLBACK = {
    "budauth": "https://www.whitehouse.gov/wp-content/uploads/2026/04/budauth_fy2027.xlsx",
    "outlays": "https://www.whitehouse.gov/wp-content/uploads/2026/04/outlays_fy2027.xlsx",
}
USA = "https://api.usaspending.gov/api/v2"
USA_TOPTIER = f"{USA}/references/toptier_agencies/"
USA_DEFC = f"{USA}/references/def_codes/"
USA_TOTAL = f"{USA}/references/total_budgetary_resources/"
USA_SUBMISSION_PERIODS = f"{USA}/references/submission_periods/"
USA_DOWNLOAD_ACCOUNTS = f"{USA}/download/accounts/"
USA_FEDERAL_ACCOUNTS = f"{USA}/federal_accounts/"

# OMB workbook layout is table-specific and mapped BY COLUMN NAME (never position), because
# the two tables differ: budauth carries 12 dimension columns + years 1976..2031 (57 year
# columns); outlays carries 13 dimension columns — an extra "Grant/non-grant split" between
# "BEA Category" and "On- or Off- Budget" — + years 1962..2031 (71 year columns, verified in
# the smoke run 2026-07-27). The melt detects the year block dynamically (first 4-digit-year
# or 'TQ' column) and maps every known dimension by name, landing the extra split losslessly in
# grant_split (NULL when absent). This is the same wide-melt format, not an improvised parser.
OMB_DIM_MAP = {
    "Agency Code": "agency_code", "Agency Name": "agency_name",
    "Bureau Code": "bureau_code", "Bureau Name": "bureau_name",
    "Account Code": "account_code", "Account Name": "account_name",
    "Treasury Agency Code": "treasury_agency_code", "CGAC Agency Code": "cgac_agency_code",
    "Subfunction Code": "subfunction_code", "Subfunction Title": "subfunction_title",
    "BEA Category": "bea_category", "On- or Off- Budget": "budget_status",
    "Grant/non-grant split": "grant_split",
}
OMB_REQUIRED_DIMS = [
    "Agency Code", "Agency Name", "Bureau Code", "Bureau Name", "Account Code",
    "Account Name", "Treasury Agency Code", "CGAC Agency Code", "Subfunction Code",
    "Subfunction Title", "BEA Category", "On- or Off- Budget",
]
OMB_ALL_FIELDS = list(OMB_DIM_MAP.values())  # incl. grant_split

# File-A (account_balances) money columns → DOUBLE; everything else stays VARCHAR.
FILE_A_MONEY_COLS = [
    "budget_authority_unobligated_balance_brought_forward",
    "adjustments_to_unobligated_balance_brought_forward_cpe",
    "budget_authority_appropriated_amount", "borrowing_authority_amount",
    "contract_authority_amount", "spending_authority_from_offsetting_collections_amount",
    "total_other_budgetary_resources_amount", "total_budgetary_resources",
    "obligations_incurred", "deobligations_or_recoveries_or_refunds_from_prior_year",
    "unobligated_balance", "gross_outlay_amount",
]
# File-A federal column names, IN ORDER — the 21-name required prefix, VERIFIED LIVE in the
# smoke run 2026-07-27. NOTE: the directive transcribed column 21 as "status"; the real column
# is "status_of_budgetary_resources_total", and the file appends one trailing provenance column
# "last_modified_date" (22 total). The gate asserts these 21 names as an in-order prefix and
# tolerates exactly one extra trailing column, printing the observed header on any drift.
FILE_A_FEDERAL_HEADER = [
    "owning_agency_name", "reporting_agency_name", "submission_period",
    "federal_account_symbol", "federal_account_name", "agency_identifier_name",
    "budget_function", "budget_subfunction",
    "budget_authority_unobligated_balance_brought_forward",
    "adjustments_to_unobligated_balance_brought_forward_cpe",
    "budget_authority_appropriated_amount", "borrowing_authority_amount",
    "contract_authority_amount", "spending_authority_from_offsetting_collections_amount",
    "total_other_budgetary_resources_amount", "total_budgetary_resources",
    "obligations_incurred", "deobligations_or_recoveries_or_refunds_from_prior_year",
    "unobligated_balance", "gross_outlay_amount", "status_of_budgetary_resources_total",
]

ACCOUNT_SWEEP_START_FY = 2017


# ── cache / governor plumbing ──────────────────────────────────────────────────────
def _cache_dir() -> str:
    d = os.environ.get("FED_APPROP_CACHE") or os.path.join(tempfile.gettempdir(),
                                                            "fed_approp_cache")
    os.makedirs(d, exist_ok=True)
    return d


def _cache_read(name: str) -> bytes | None:
    p = os.path.join(_cache_dir(), name)
    if os.path.exists(p) and os.path.getsize(p) > 0:
        with open(p, "rb") as fh:
            return fh.read()
    return None


def _cache_write(name: str, data: bytes) -> None:
    p = os.path.join(_cache_dir(), name)
    tmp = p + ".tmp"
    with open(tmp, "wb") as fh:
        fh.write(data)
    os.replace(tmp, p)


_GOVS: dict[str, RateGovernor] = {}


def _gov(url_or_host: str) -> RateGovernor:
    host = url_or_host
    if "://" in host:
        host = host.split("://", 1)[1].split("/", 1)[0]
    if host not in _GOVS:
        ckpt = os.path.join(_cache_dir(), f"gov_{host.replace('.', '_')}.ckpt.json")
        _GOVS[host] = RateGovernor(host=host, checkpoint_path=ckpt)
    return _GOVS[host]


def _get_json(url: str, *, ok404: bool = False) -> dict | list | None:
    gov = _gov(url)
    resp = gov.get(url)
    code = getattr(resp, "status_code", None)
    if code == 200:
        return resp.json()
    if code == 404 and ok404:
        return None
    raise RuntimeError(f"GET {url} -> HTTP {code}")


def _post_json(url: str, body: dict) -> dict:
    gov = _gov(url)
    resp = gov.post(url, json_body=body)
    code = getattr(resp, "status_code", None)
    if code != 200:
        raise RuntimeError(f"POST {url} -> HTTP {code}: {(getattr(resp, 'text', '') or '')[:200]}")
    return resp.json()


def _get_bytes(url: str) -> bytes:
    gov = _gov(url)
    resp = gov.get(url)
    code = getattr(resp, "status_code", None)
    if code != 200:
        raise RuntimeError(f"GET {url} -> HTTP {code}")
    return resp.content


# ── coercion ────────────────────────────────────────────────────────────────────────
def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if not s or s in {"", "-", "--", "NA", "N/A", "n.a."}:
        return None
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _pipe(v) -> str | None:
    """L54 — no LIST<VARCHAR> in Lance 1.5.x; multi-value → pipe-joined VARCHAR."""
    if v is None:
        return None
    if isinstance(v, list):
        return "|".join(str(x) for x in v) if v else None
    if isinstance(v, dict):
        return json.dumps(v, sort_keys=True)
    s = str(v).strip()
    return s or None


# ── ledger (ops.federal_appropriations_ingest_runs) ────────────────────────────────
LEDGER_DDL = """
CREATE SCHEMA IF NOT EXISTS ops;
CREATE TABLE IF NOT EXISTS ops.federal_appropriations_ingest_runs (
    id            bigint GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
    run_id        uuid        NOT NULL,
    stream        text        NOT NULL,
    resolved_url  text,
    source_bytes  bigint,
    rows_written  bigint,
    datasets      jsonb,
    started_at    timestamptz,
    finished_at   timestamptz,
    status        text        NOT NULL
                              CHECK (status IN ('running', 'completed', 'failed')),
    disposition   text,
    notes         text,
    recorded_at   timestamptz NOT NULL DEFAULT now()
);
CREATE INDEX IF NOT EXISTS federal_appropriations_ingest_runs_run_idx
    ON ops.federal_appropriations_ingest_runs (run_id);
CREATE INDEX IF NOT EXISTS federal_appropriations_ingest_runs_stream_idx
    ON ops.federal_appropriations_ingest_runs (stream);
CREATE INDEX IF NOT EXISTS federal_appropriations_ingest_runs_status_started_idx
    ON ops.federal_appropriations_ingest_runs (status, started_at);
"""

CATALOG_DDL = """
CREATE TABLE IF NOT EXISTS ops.data_source_catalog (
    id                          bigint GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
    source_slug                 text        NOT NULL UNIQUE,
    display_name                text        NOT NULL,
    strategic_role              text,
    r2_prefix                   text,
    refresh_cadence             text,
    lifecycle_stage             text,
    owner_team                  text,
    is_active                   boolean     NOT NULL DEFAULT true,
    audit_ledger_table          text,
    essentials_mv_name          text,
    source_url                  text,
    notes                       text,
    bridge_source_name_patterns text,
    audience_mv_name_patterns   text,
    created_at                  timestamptz NOT NULL DEFAULT now()
);
"""

CATALOG_ROWS = [
    dict(
        source_slug="omb_public_budget_database",
        display_name="OMB Public Budget Database (Budget Authority + Outlays)",
        strategic_role="Appropriated budget authority + outlays at account×year grain — the "
                       "appropriated side of the federal dollar (1976–2031, actuals + vintage estimates).",
        r2_prefix="active/omb_budget_authority/",
        refresh_cadence="annual (Feb–Apr budget release)",
        lifecycle_stage="active",
        owner_team="data-factory",
        is_active=True,
        audit_ledger_table="ops.federal_appropriations_ingest_runs",
        source_url=OMB_SUPPLEMENTAL,
        notes="budauth_fy{YYYY}.xlsx + outlays_fy{YYYY}.xlsx; also lands active/omb_outlays/.",
    ),
    dict(
        source_slug="usaspending_budgetary_resources",
        display_name="USAspending Budgetary Resources & Account Balances (File A)",
        strategic_role="Appropriated / obligated / unobligated at federal-account, TAS and "
                       "agency×FY grain (File A account balances + agency budgetary resources).",
        r2_prefix="active/usaspending_federal_account_balances/",
        refresh_cadence="quarterly (DABS submissions)",
        lifecycle_stage="active",
        owner_team="data-factory",
        is_active=True,
        audit_ledger_table="ops.federal_appropriations_ingest_runs",
        source_url="https://api.usaspending.gov/",
        notes="download/accounts File A (federal+treasury) + agency budgetary_resources + "
              "def_codes + total_budgetary_resources + federal_accounts roster.",
    ),
]


def _dsn() -> str | None:
    dsn = os.environ.get("HQX_DB_URL_POOLED")
    if not dsn:
        print("WARN: HQX_DB_URL_POOLED not set; skipping HQX writes.", flush=True)
    return dsn


def apply_state_schema() -> None:
    """Apply the ledger + catalog DDL and register the sources (idempotent)."""
    dsn = _dsn()
    if not dsn:
        raise RuntimeError("HQX_DB_URL_POOLED not set (Doppler core-x/prd).")
    import psycopg

    with psycopg.connect(dsn, autocommit=True) as conn, conn.cursor() as cur:
        cur.execute(LEDGER_DDL)
        cur.execute(CATALOG_DDL)
        _register_sources(cur)
    print("Applied ops.federal_appropriations_ingest_runs + registered sources.", flush=True)


def _register_sources(cur) -> None:
    cols = ["source_slug", "display_name", "strategic_role", "r2_prefix", "refresh_cadence",
            "lifecycle_stage", "owner_team", "is_active", "audit_ledger_table", "source_url",
            "notes"]
    for row in CATALOG_ROWS:
        cur.execute(
            f"INSERT INTO ops.data_source_catalog ({', '.join(cols)}) "
            f"VALUES ({', '.join(['%s'] * len(cols))}) "
            f"ON CONFLICT (source_slug) DO NOTHING",
            tuple(row[c] for c in cols),
        )


def _record_run(run_id, stream, resolved_url, source_bytes, rows_written, datasets,
                started_at, finished_at, status, disposition, notes) -> None:
    dsn = _dsn()
    if not dsn:
        return
    try:
        import psycopg

        with psycopg.connect(dsn, autocommit=True) as conn, conn.cursor() as cur:
            cur.execute(LEDGER_DDL)
            cur.execute(
                """
                INSERT INTO ops.federal_appropriations_ingest_runs
                    (run_id, stream, resolved_url, source_bytes, rows_written, datasets,
                     started_at, finished_at, status, disposition, notes)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (str(run_id), stream, resolved_url, source_bytes, rows_written,
                 json.dumps(datasets), started_at, finished_at, status, disposition,
                 (notes or "")[:8000]),
            )
    except Exception as exc:  # noqa: BLE001 — audit must never mask a good load
        print(f"WARN: ledger write failed: {exc}", flush=True)


def _write_lance(table, uri: str, so: dict, *, mode: str = "overwrite") -> None:
    import lance

    lance.write_dataset(table, uri, mode=mode, data_storage_version=DATA_STORAGE_VERSION,
                        max_rows_per_file=MAX_ROWS_PER_FILE, max_bytes_per_file=MAX_BYTES_PER_FILE,
                        storage_options=so)


# ═══════════════════════════════════════════════════════════════════════════════════
# OMB Public Budget Database — shared workbook melt (budget authority + outlays)
# ═══════════════════════════════════════════════════════════════════════════════════
def _resolve_omb_urls() -> dict[str, str]:
    """Discover the current budauth/outlays xlsx off the supplemental-materials page
    (§2.3 — release-dated URLs move every year). Falls back to the verified pair."""
    try:
        html = _get_bytes(OMB_SUPPLEMENTAL).decode("utf-8", "replace")
    except ThrottleHalt:
        raise  # a breaker halt on the discovery page must stop the run, not fetch more
    except Exception as exc:  # noqa: BLE001
        print(f"WARN: OMB discovery page fetch failed ({exc}); using verified fallback.", flush=True)
        return dict(OMB_FALLBACK)
    out: dict[str, str] = {}
    for m in re.finditer(r'href="([^"]*(budauth|outlays)_fy\d{4}\.xlsx)"', html, re.I):
        href = m.group(1)
        if not href.startswith("http"):
            href = "https://www.whitehouse.gov" + (href if href.startswith("/") else "/" + href)
        out.setdefault(m.group(2).lower(), href)
    if "budauth" not in out or "outlays" not in out:
        print(f"WARN: OMB discovery incomplete {out}; filling from verified fallback.", flush=True)
        for k, v in OMB_FALLBACK.items():
            out.setdefault(k, v)
    return out


def _omb_schema():
    import pyarrow as pa

    return pa.schema([
        ("agency_code", pa.string()), ("agency_name", pa.string()),
        ("bureau_code", pa.string()), ("bureau_name", pa.string()),
        ("account_code", pa.string()), ("account_name", pa.string()),
        ("treasury_agency_code", pa.string()), ("cgac_agency_code", pa.string()),
        ("subfunction_code", pa.string()), ("subfunction_title", pa.string()),
        ("bea_category", pa.string()), ("budget_status", pa.string()),
        ("grant_split", pa.string()),   # populated only for outlays (extra dimension)
        ("year", pa.int32()), ("is_transition_quarter", pa.bool_()),
        ("is_estimate", pa.bool_()), ("budget_year", pa.int32()),
        ("value_musd", pa.float64()), ("period_label", pa.string()),
        ("row_kind", pa.string()), ("source", pa.string()),
        ("ingested_at", pa.timestamp("us", tz="UTC")),
    ])


def _is_year_label(lab) -> bool:
    return lab == "TQ" or (bool(lab) and lab.isdigit() and len(lab) == 4)


def _melt_omb(raw: bytes, filename: str, source_url: str, *, smoke: bool):
    """Wide→long melt of an OMB workbook, layout-adaptive by COLUMN NAME. Performs the §2.2 L44
    discovery print (sheetnames, dims, header) BEFORE parsing; hard-fails only if a REQUIRED
    dimension name is missing (real drift), not on a known extra dim / wider year range."""
    import openpyxl

    m = re.search(r"(?:budauth|outlays)_fy(\d{4})", filename, re.I)
    if not m:
        raise RuntimeError(f"cannot parse budget_year vintage from filename {filename!r}")
    budget_year = int(m.group(1))

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheetnames = list(wb.sheetnames)
        ws = wb[sheetnames[0]]
        max_row, max_col = ws.max_row, ws.max_column
        it = ws.iter_rows(values_only=True)
        header = [_s(h) for h in next(it)]
        # ── L44 discovery (mandatory for outlays; harmless for budauth) ──
        print(f"[omb] file={filename} budget_year={budget_year} sheets={sheetnames} "
              f"dims={max_row}x{max_col}", flush=True)
        print(f"[omb] full header={header}", flush=True)

        # the year block is everything from the first 4-digit-year / 'TQ' column onward
        year_start = next((j for j, lab in enumerate(header) if _is_year_label(lab)), None)
        if year_start is None:
            raise RuntimeError(f"OMB: no year columns found; header={header}")
        dim_labels = header[:year_start]
        dim_cols = []  # (col_idx, field)
        present = set()
        for j, lab in enumerate(dim_labels):
            field = OMB_DIM_MAP.get(lab)
            if field:
                dim_cols.append((j, field))
                present.add(lab)
            else:
                print(f"  WARN OMB unknown dimension column {lab!r} at col {j} — skipped", flush=True)
        missing = [d for d in OMB_REQUIRED_DIMS if d not in present]
        if missing:
            raise RuntimeError(f"OMB missing required dimension(s) {missing}; observed dims "
                               f"{dim_labels} (full header: {header})")

        year_cols = []  # (col_idx, year_int, is_tq, period_label)
        for j in range(year_start, len(header)):
            lab = header[j]
            if lab == "TQ":
                year_cols.append((j, -1976, True, "TQ1976"))
            elif _is_year_label(lab):
                year_cols.append((j, int(lab), False, f"FY{int(lab)}"))
            else:
                print(f"  WARN OMB non-year column {lab!r} at col {j} in year block — skipped",
                      flush=True)

        rows: list[dict] = []
        ingested_at = dt.datetime.now(dt.timezone.utc)
        n = 0
        for r in it:
            n += 1
            if smoke and n > 200:
                break
            base = {f: None for f in OMB_ALL_FIELDS}
            for (j, field) in dim_cols:
                base[field] = _s(r[j]) if j < len(r) else None
            row_kind = "account" if base.get("account_code") else "aggregate_receipt"
            for (j, yr, is_tq, plabel) in year_cols:
                if j >= len(r):
                    continue
                v = _num(r[j])
                if v is None:
                    continue
                rec = dict(base)
                rec.update({
                    "year": yr, "is_transition_quarter": is_tq,
                    "is_estimate": (not is_tq) and (yr >= budget_year - 1),
                    "budget_year": budget_year,
                    # UNIT CORRECTION: the OMB Public Budget Database DETAIL files (budauth/
                    # outlays xlsx) are published in THOUSANDS of dollars, NOT millions as the
                    # directive stated (verified 2026-07-27: Social Security OASI FY2025 =
                    # 1,430,282,000 thousand = $1.43T; net BA = $7.6T; discretionary = $1.89T).
                    # Convert thousands → millions so value_musd is truthfully $millions and the
                    # derived reconciliation (omb_*_musd × 1e6 ≈ USAspending $usd) is meaningful.
                    "value_musd": v / 1000.0, "period_label": plabel,
                    "row_kind": row_kind, "source": source_url, "ingested_at": ingested_at,
                })
                rows.append(rec)
    finally:
        wb.close()
    layout = {"sheets": sheetnames, "dims": [max_row, max_col], "n_dims": len(dim_cols),
              "n_year_cols": len(year_cols), "year_min": min(header[year_start:], default=None),
              "budget_year": budget_year, "full_header": header}
    return rows, budget_year, layout


def _gate_omb(rows, budget_year, layout, *, is_budauth: bool):
    years = {r["year"] for r in rows}
    assert 1976 in years and 2031 in years, f"year span missing 1976/2031: {sorted(years)[:3]}…"
    n = len(rows)
    if is_budauth:
        # budauth: exactly 57 year columns (1976..2031 + TQ), rows in the directive band
        assert layout["n_year_cols"] == 57, \
            f"budauth year-col count {layout['n_year_cols']} != 57 (schema drift)"
        assert 200_000 <= n <= 400_000, f"budauth row count {n:,} outside [200K, 400K]"
    else:
        # outlays: verified 71 year columns (1962..2031 + TQ) and 13 dims — a wider, well-formed
        # variant; require the superset year block and a defensible row floor, print observed.
        assert layout["n_year_cols"] >= 57, \
            f"outlays year-col count {layout['n_year_cols']} < 57 (schema drift)"
        assert 150_000 <= n <= 600_000, f"outlays row count {n:,} outside [150K, 600K]"
    est_years = {r["year"] for r in rows if r["is_estimate"]}
    for y in years:
        if y == -1976:
            continue
        should = y >= budget_year - 1
        assert (y in est_years) == should, \
            f"is_estimate wrong for FY{y} (budget_year={budget_year}, should_estimate={should})"
    gross = sum(r["value_musd"] for r in rows
                if r["year"] == 2025 and r["row_kind"] == "account" and r["value_musd"] > 0)
    offset = sum(r["value_musd"] for r in rows if r["year"] == 2025 and r["value_musd"] < 0)
    assert offset != 0, "FY2025 offsetting/receipt rows summed to 0 — negatives were dropped, not summed"
    if is_budauth:
        assert 1_000_000 <= gross <= 30_000_000, \
            f"FY2025 gross BA {gross:,.0f} musd outside defensible band [1M, 30M]"
    else:
        assert gross > 0, f"FY2025 gross outlays {gross:,.0f} musd not positive"


def _run_omb(stream: str, uri: str, source_url: str, *, so: dict, smoke: bool, run_id) -> dict:
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    status, disp, notes, built, src_bytes = "failed", "none", "", [], None
    resolved = source_url
    try:
        is_budauth = stream == "omb_budauth"
        # cache the ~1.6–2.1 MB workbook so a re-run does not re-fetch whitehouse.gov
        cache_name = os.path.basename(resolved)
        raw = _cache_read(cache_name)
        if raw is None:
            raw = _get_bytes(resolved)
            _cache_write(cache_name, raw)
        src_bytes = len(raw)
        print(f"[{stream}] {resolved} -> {src_bytes:,} bytes", flush=True)

        rows, budget_year, layout = _melt_omb(raw, cache_name, resolved, smoke=smoke)
        compact = {k: layout[k] for k in ("sheets", "dims", "n_dims", "n_year_cols",
                                          "year_min", "budget_year")}
        notes = f"budget_year={budget_year}; layout={json.dumps(compact)}"
        tbl = pa.Table.from_pylist(rows, schema=_omb_schema())
        _write_lance(tbl, uri, so)
        built = _build_indexes(uri, btree=["agency_code", "account_code", "year"],
                               bitmap=[], so=so)
        _gov(resolved).mark_done(f"omb:{cache_name}")
        if not smoke:
            _gate_omb(rows, budget_year, layout, is_budauth=is_budauth)
        status, disp = "completed", "ok"
    except ThrottleHalt:
        disp = "throttled"
        raise
    except Exception as exc:  # noqa: BLE001
        notes = f"{notes} ERROR: {exc}"[:8000]
        raise
    finally:
        _record_run(run_id, stream, resolved, src_bytes, len(rows),
                    {os.path.basename(uri.rstrip('/')): len(rows)}, t0,
                    dt.datetime.now(dt.timezone.utc), status, disp, notes)
        print(f"[{stream}] rows={len(rows):,} status={status} disp={disp}", flush=True)
    return {"stream": stream, "rows": len(rows), "status": status, "uri": uri}


# ═══════════════════════════════════════════════════════════════════════════════════
# USAspending — agency budgetary resources (drives usa_agency #5 + usa_agency_periods #6)
# ═══════════════════════════════════════════════════════════════════════════════════
def _toptier_roster() -> list[dict]:
    data = _get_json(USA_TOPTIER)
    results = data.get("results", data) if isinstance(data, dict) else data
    return list(results or [])


def _agency_budgetary(code: str) -> dict:
    """Per-agency budgetary_resources, cached by toptier_code so #5 and #6 fetch once."""
    cache_name = f"agency_budg_{code}.json"
    gov = _gov(USA_TOPTIER)
    cached = _cache_read(cache_name)
    if cached is not None:
        try:
            parsed = json.loads(cached)
            gov.mark_done(f"agency:{code}")
            return parsed
        except ValueError:
            pass
    data = _get_json(f"{USA}/agency/{code}/budgetary_resources/")
    _cache_write(cache_name, json.dumps(data).encode("utf-8"))
    gov.mark_done(f"agency:{code}")
    return data


def _sweep_agencies(*, smoke: bool):
    roster = _toptier_roster()
    if smoke:
        roster = roster[:3]
    name_by_code, out = {}, []
    skipped = []
    for a in roster:
        code = _s(a.get("toptier_code"))
        if not code:
            continue
        name_by_code[code] = a
        try:
            out.append((code, _agency_budgetary(code)))
        except ThrottleHalt:
            raise
        except Exception as exc:  # noqa: BLE001 — per-agency failure is logged, never fatal
            skipped.append({"code": code, "error": str(exc)[:200]})
            print(f"  WARN agency {code} skipped: {exc}", flush=True)
    return name_by_code, out, skipped


def run_usa_agency(*, so: dict, uri: str, smoke: bool, run_id) -> dict:
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    status, disp, notes, built = "failed", "none", "", []
    try:
        name_by_code, swept, skipped = _sweep_agencies(smoke=smoke)
        ingested_at = dt.datetime.now(dt.timezone.utc)
        for code, data in swept:
            meta = name_by_code.get(code, {})
            for y in (data.get("agency_data_by_year") or []):
                res = y.get("agency_budgetary_resources")
                if res is None:
                    continue  # gate: every landed agency_budgetary_resources is non-null
                rows.append({
                    "toptier_code": code,
                    "agency_name": _s(meta.get("agency_name")),
                    "abbreviation": _s(meta.get("abbreviation")),
                    "agency_slug": _s(meta.get("agency_slug")),
                    "fiscal_year": int(y["fiscal_year"]),
                    "agency_budgetary_resources": float(res),
                    "agency_total_obligated": _num(y.get("agency_total_obligated")),
                    "agency_total_outlayed": _num(y.get("agency_total_outlayed")),
                    "total_budgetary_resources": _num(y.get("total_budgetary_resources")),
                    "source": USA_TOPTIER, "ingested_at": ingested_at,
                })
        schema = pa.schema([
            ("toptier_code", pa.string()), ("agency_name", pa.string()),
            ("abbreviation", pa.string()), ("agency_slug", pa.string()),
            ("fiscal_year", pa.int32()), ("agency_budgetary_resources", pa.float64()),
            ("agency_total_obligated", pa.float64()), ("agency_total_outlayed", pa.float64()),
            ("total_budgetary_resources", pa.float64()), ("source", pa.string()),
            ("ingested_at", pa.timestamp("us", tz="UTC")),
        ])
        _write_lance(pa.Table.from_pylist(rows, schema=schema), uri, so)
        built = _build_indexes(uri, btree=["toptier_code", "fiscal_year"], bitmap=[], so=so)
        n_agencies = len({r["toptier_code"] for r in rows})
        notes = f"agencies={n_agencies}; skipped={json.dumps(skipped)[:1000]}"
        if not smoke:
            assert n_agencies >= 100, f"only {n_agencies} agencies (<100)"
            assert all(r["agency_budgetary_resources"] is not None for r in rows)
            gw2025 = [r["total_budgetary_resources"] for r in rows
                      if r["fiscal_year"] == 2025 and r["total_budgetary_resources"]]
            assert gw2025 and max(gw2025) > 10_000_000_000_000, \
                f"FY2025 gov-wide total_budgetary_resources not > $10T: {max(gw2025) if gw2025 else None}"
        status, disp = "completed", ("partial" if skipped else "ok")
    except ThrottleHalt:
        disp = "throttled"
        raise
    except Exception as exc:  # noqa: BLE001
        notes = f"{notes} ERROR: {exc}"[:8000]
        raise
    finally:
        _record_run(run_id, "usa_agency", USA_TOPTIER, None, len(rows),
                    {"usaspending_agency_fy": len(rows)}, t0,
                    dt.datetime.now(dt.timezone.utc), status, disp, notes)
        print(f"[usa_agency] rows={len(rows):,} status={status} disp={disp}", flush=True)
    return {"stream": "usa_agency", "rows": len(rows), "status": status, "uri": uri}


def run_usa_agency_periods(*, so: dict, uri: str, smoke: bool, run_id) -> dict:
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    status, disp, notes, built = "failed", "none", "", []
    try:
        _, swept, skipped = _sweep_agencies(smoke=smoke)
        ingested_at = dt.datetime.now(dt.timezone.utc)
        for code, data in swept:
            for y in (data.get("agency_data_by_year") or []):
                fy = int(y["fiscal_year"])
                for p in (y.get("agency_obligation_by_period") or []):
                    if p.get("period") is None:
                        continue
                    rows.append({
                        "toptier_code": code, "fiscal_year": fy,
                        "period": int(p["period"]), "obligated": _num(p.get("obligated")),
                        "source": USA_TOPTIER, "ingested_at": ingested_at,
                    })
        schema = pa.schema([
            ("toptier_code", pa.string()), ("fiscal_year", pa.int32()),
            ("period", pa.int32()), ("obligated", pa.float64()),
            ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
        ])
        _write_lance(pa.Table.from_pylist(rows, schema=schema), uri, so)
        built = _build_indexes(uri, btree=["toptier_code", "fiscal_year", "period"],
                               bitmap=[], so=so)
        notes = f"skipped={json.dumps(skipped)[:1000]}"
        status, disp = "completed", ("partial" if skipped else "ok")
    except ThrottleHalt:
        disp = "throttled"
        raise
    except Exception as exc:  # noqa: BLE001
        notes = f"{notes} ERROR: {exc}"[:8000]
        raise
    finally:
        _record_run(run_id, "usa_agency_periods", USA_TOPTIER, None, len(rows),
                    {"usaspending_agency_period_obligations": len(rows)}, t0,
                    dt.datetime.now(dt.timezone.utc), status, disp, notes)
        print(f"[usa_agency_periods] rows={len(rows):,} status={status} disp={disp}", flush=True)
    return {"stream": "usa_agency_periods", "rows": len(rows), "status": status, "uri": uri}


# ═══════════════════════════════════════════════════════════════════════════════════
# USAspending — total budgetary resources (#7), DEF codes (#9), account roster (#8)
# ═══════════════════════════════════════════════════════════════════════════════════
def run_usa_total_resources(*, so: dict, uri: str, smoke: bool, run_id) -> dict:
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    status, disp, notes, built = "failed", "none", "", []
    try:
        data = _get_json(USA_TOTAL)
        results = data.get("results", data) if isinstance(data, dict) else data
        ingested_at = dt.datetime.now(dt.timezone.utc)
        for r in (results or []):
            rows.append({
                "fiscal_year": int(r["fiscal_year"]),
                "fiscal_period": int(r["fiscal_period"]),
                "total_budgetary_resources": _num(r.get("total_budgetary_resources")),
                "source": USA_TOTAL, "ingested_at": ingested_at,
            })
        schema = pa.schema([
            ("fiscal_year", pa.int32()), ("fiscal_period", pa.int32()),
            ("total_budgetary_resources", pa.float64()),
            ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
        ])
        _write_lance(pa.Table.from_pylist(rows, schema=schema), uri, so)
        built = _build_indexes(uri, btree=["fiscal_year", "fiscal_period"], bitmap=[], so=so)
        if not smoke:
            assert len(rows) >= 100, f"only {len(rows)} total-resource rows (<100)"
        status, disp = "completed", "ok"
    except ThrottleHalt:
        disp = "throttled"
        raise
    except Exception as exc:  # noqa: BLE001
        notes = f"{notes} ERROR: {exc}"[:8000]
        raise
    finally:
        _record_run(run_id, "usa_total_resources", USA_TOTAL, None, len(rows),
                    {"usaspending_total_budgetary_resources": len(rows)}, t0,
                    dt.datetime.now(dt.timezone.utc), status, disp, notes)
        print(f"[usa_total_resources] rows={len(rows):,} status={status} disp={disp}", flush=True)
    return {"stream": "usa_total_resources", "rows": len(rows), "status": status, "uri": uri}


def run_usa_defc(*, so: dict, uri: str, smoke: bool, run_id) -> dict:
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    status, disp, notes, built = "failed", "none", "", []
    try:
        data = _get_json(USA_DEFC)
        results = data.get("codes", data.get("results", data)) if isinstance(data, dict) else data
        ingested_at = dt.datetime.now(dt.timezone.utc)
        all_keys: set[str] = set()
        for r in (results or []):
            all_keys.update(r.keys())
        all_keys = sorted(all_keys)
        for r in (results or []):
            rec = {k: _pipe(r.get(k)) for k in all_keys}
            rec["source"] = USA_DEFC
            rec["ingested_at"] = ingested_at
            rows.append(rec)
        fields = [(k, pa.string()) for k in all_keys]
        fields += [("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC"))]
        _write_lance(pa.Table.from_pylist(rows, schema=pa.schema(fields)), uri, so)
        built = _build_indexes(uri, btree=["code"], bitmap=[], so=so)
        codes = {r.get("code") for r in rows}
        notes = f"defc_count={len(rows)} (baseline 52); has_Z={'Z' in codes}"
        assert "Z" in codes, "DEFC 'Z' (IIJA, P.L. 117-58) absent — hard fail"
        if len(rows) != 52:
            notes += f" | WARN registry drift: {len(rows)} != 52 baseline"
            print(f"  WARN: DEFC count {len(rows)} != 52 baseline", flush=True)
        status, disp = "completed", "ok"
    except ThrottleHalt:
        disp = "throttled"
        raise
    except Exception as exc:  # noqa: BLE001
        notes = f"{notes} ERROR: {exc}"[:8000]
        raise
    finally:
        _record_run(run_id, "usa_defc", USA_DEFC, None, len(rows),
                    {"usaspending_def_codes": len(rows)}, t0,
                    dt.datetime.now(dt.timezone.utc), status, disp, notes)
        print(f"[usa_defc] rows={len(rows):,} status={status} disp={disp}", flush=True)
    return {"stream": "usa_defc", "rows": len(rows), "status": status, "uri": uri}


def run_usa_account_roster(*, so: dict, uri: str, smoke: bool, run_id) -> dict:
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    status, disp, notes, built = "failed", "none", "", []
    fy = _current_fiscal_year()
    try:
        ingested_at = dt.datetime.now(dt.timezone.utc)
        page, seen = 1, set()
        while True:
            data = _post_json(USA_FEDERAL_ACCOUNTS, {"fy": str(fy), "limit": 100, "page": page})
            results = data.get("results", []) if isinstance(data, dict) else (data or [])
            for r in results:
                acct = _s(r.get("account_number"))
                if not acct or acct in seen:
                    continue
                seen.add(acct)
                aid = _s(r.get("agency_identifier"))
                acode = acct.split("-", 1)[1] if "-" in acct else None
                rows.append({
                    "account_number": acct, "agency_identifier": aid, "account_code": acode,
                    "account_name": _s(r.get("account_name")),
                    "account_id": int(r["account_id"]) if r.get("account_id") is not None else None,
                    "managing_agency": _s(r.get("managing_agency")),
                    "managing_agency_acronym": _s(r.get("managing_agency_acronym")),
                    "source": USA_FEDERAL_ACCOUNTS, "ingested_at": ingested_at,
                })
            has_next = data.get("hasNext") if isinstance(data, dict) else False
            if smoke or not has_next or not results:
                break
            page += 1
        schema = pa.schema([
            ("account_number", pa.string()), ("agency_identifier", pa.string()),
            ("account_code", pa.string()), ("account_name", pa.string()),
            ("account_id", pa.int64()), ("managing_agency", pa.string()),
            ("managing_agency_acronym", pa.string()),
            ("source", pa.string()), ("ingested_at", pa.timestamp("us", tz="UTC")),
        ])
        _write_lance(pa.Table.from_pylist(rows, schema=schema), uri, so)
        built = _build_indexes(uri, btree=["account_number"], bitmap=[], so=so)
        notes = f"fy={fy}; pages={page}"
        status, disp = "completed", "ok"
    except ThrottleHalt:
        disp = "throttled"
        raise
    except Exception as exc:  # noqa: BLE001
        notes = f"{notes} ERROR: {exc}"[:8000]
        raise
    finally:
        _record_run(run_id, "usa_account_roster", USA_FEDERAL_ACCOUNTS, None, len(rows),
                    {"usaspending_federal_account_roster": len(rows)}, t0,
                    dt.datetime.now(dt.timezone.utc), status, disp, notes)
        print(f"[usa_account_roster] rows={len(rows):,} status={status} disp={disp}", flush=True)
    return {"stream": "usa_account_roster", "rows": len(rows), "status": status, "uri": uri}


# ═══════════════════════════════════════════════════════════════════════════════════
# USAspending — File-A account balances (#3 federal + #4 treasury) via download/accounts
# ═══════════════════════════════════════════════════════════════════════════════════
def _current_fiscal_year() -> int:
    now = dt.datetime.now(dt.timezone.utc)
    return now.year + 1 if now.month >= 10 else now.year


def _resolve_account_sweep(*, smoke: bool) -> list[tuple[int, int]]:
    """(fy, quarter) pairs to mint: FY2017→latest, quarter 4 for closed years and the latest
    available quarter for the open year (probed from submission_periods)."""
    try:
        data = _get_json(USA_SUBMISSION_PERIODS)
        periods = (data.get("available_periods") or data.get("results") or []) \
            if isinstance(data, dict) else (data or [])
    except ThrottleHalt:
        raise  # a breaker halt must stop the run, not fall through to mint the sweep
    except Exception as exc:  # noqa: BLE001
        print(f"WARN: submission_periods probe failed ({exc}); default FY2017→2025 Q4.", flush=True)
        periods = []
    by_fy: dict[int, set] = {}
    for p in periods:
        fy = p.get("submission_fiscal_year")
        q = p.get("submission_fiscal_quarter")
        if fy is None or q is None:
            continue
        by_fy.setdefault(int(fy), set()).add(int(q))
    if by_fy:
        max_fy = max(by_fy)
        sweep = []
        for fy in range(ACCOUNT_SWEEP_START_FY, max_fy + 1):
            qs = by_fy.get(fy, set())
            sweep.append((fy, 4 if 4 in qs else (max(qs) if qs else 4)))
    else:
        sweep = [(fy, 4) for fy in range(ACCOUNT_SWEEP_START_FY, _current_fiscal_year())]
    if smoke:
        sweep = sweep[-1:]
    return sweep


def _mint_and_download(account_level: str, fy: int, quarter: int) -> bytes:
    """Return the File-A ZIP for (level, fy, quarter). The DETERMINISTIC on-disk cache is
    checked BEFORE the mint POST (keyed by level/fy/quarter, not the mint-returned filename),
    so a warm restart after a throttle-halt re-fetches nothing and NEVER re-mints against a
    host that just blocked us. Completed units are recorded in the governor checkpoint."""
    gov = _gov(USA_DOWNLOAD_ACCOUNTS)
    ckpt_key = f"fileA:{account_level}:{fy}:q{quarter}"
    det_cache = f"fileA_{account_level}_fy{fy}_q{quarter}.zip"
    cached = _cache_read(det_cache)
    if cached is not None:
        print(f"    cached {det_cache} ({len(cached):,} B) — skip mint", flush=True)
        gov.mark_done(ckpt_key)
        return cached
    body = {"account_level": account_level,
            "filters": {"fy": fy, "quarter": quarter, "submission_types": ["account_balances"]},
            "file_format": "csv"}
    mint = _post_json(USA_DOWNLOAD_ACCOUNTS, body)
    status_url = mint["status_url"]
    file_url = mint["file_url"]
    for _ in range(40):  # poll to finished (~10 s cadence, cap 40)
        st = _get_json(status_url)
        state = (st or {}).get("status")
        if state == "finished":
            break
        if state == "failed":
            raise RuntimeError(f"download job failed fy={fy} level={account_level}: {st}")
        time.sleep(10)
    else:
        raise TimeoutError(f"download job did not finish fy={fy} level={account_level} (40 polls)")
    data = _get_bytes(file_url)
    _cache_write(det_cache, data)
    gov.mark_done(ckpt_key)
    return data


def _parse_file_a(zip_bytes: bytes, fy: int, account_level_label: str):
    """Extract the single CSV from the File-A ZIP and project via DuckDB (money→DOUBLE,
    everything else verbatim VARCHAR + provenance). Returns (arrow_table, header_list)."""
    import duckdb

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        members = [n for n in zf.namelist() if n.lower().endswith(".csv")]
        if not members:
            raise RuntimeError(f"no CSV in File-A ZIP fy={fy} level={account_level_label}")
        csv_bytes = zf.read(members[0])
    tmp = os.path.join(_cache_dir(), f"_fileA_{account_level_label}_{fy}.csv")
    with open(tmp, "wb") as fh:
        fh.write(csv_bytes)
    con = duckdb.connect(":memory:")
    try:
        read = (f"read_csv('{tmp}', header=true, all_varchar=true, "
                "strict_mode=false, null_padding=true, sample_size=-1)")
        header = [d[0] for d in con.execute(f"SELECT * FROM {read} LIMIT 0").description]
        proj = []
        for col in header:
            if col in FILE_A_MONEY_COLS:
                proj.append(f"TRY_CAST(replace(replace(nullif(trim(\"{col}\"),''),',',''),'$','') "
                            f"AS DOUBLE) AS \"{col}\"")
            else:
                proj.append(f"nullif(trim(\"{col}\"), '') AS \"{col}\"")
        proj.append("TRY_CAST(regexp_extract(\"submission_period\", 'FY(\\d{4})', 1) AS INTEGER) "
                    "AS fiscal_year")
        proj.append("TRY_CAST(regexp_extract(\"submission_period\", 'P(\\d{2})', 1) AS INTEGER) "
                    "AS fiscal_period")
        proj.append(f"'{account_level_label}' AS account_level")
        proj.append("? AS source")
        proj.append("now() AS ingested_at")
        sql = f"SELECT {', '.join(proj)} FROM {read}"
        tbl = con.execute(sql, [USA_DOWNLOAD_ACCOUNTS]).to_arrow_table()
    finally:
        con.close()
        try:
            os.remove(tmp)
        except OSError:
            pass
    return tbl, header


def _filea_header_ok(observed) -> bool:
    """The directive-verified names must appear in order; the live file appends at most one
    trailing provenance column (last_modified_date) → 21 or 22 total."""
    k = len(FILE_A_FEDERAL_HEADER)
    return (observed is not None and list(observed[:k]) == FILE_A_FEDERAL_HEADER
            and len(observed) in (k, k + 1))


def _gate_file_a_federal(combined, per_fy: dict, headers: dict) -> None:
    """§8 File-A gates (federal level, non-smoke): verified 22-col header in order, FY2025 row
    band, back-sweep row floor + header, and appropriated non-null > 90% on FY2025."""
    import pyarrow.compute as pc

    assert _filea_header_ok(headers.get(2025)), \
        f"FY2025 federal header drift: observed {headers.get(2025)}"
    assert 1_700 <= per_fy.get(2025, 0) <= 2_300, \
        f"FY2025 federal rows {per_fy.get(2025)} outside [1700,2300]"
    for fy, n in per_fy.items():
        if fy == 2025:
            continue
        assert n >= 1_000, f"FY{fy} federal rows {n} < 1000 (empty/malformed early-year ZIP)"
        assert _filea_header_ok(headers.get(fy)), f"FY{fy} header drift: {headers.get(fy)}"
    fy25 = combined.filter(pc.equal(combined["fiscal_year"], 2025))
    appro = fy25["budget_authority_appropriated_amount"]
    frac = (len(appro) - appro.null_count) / max(1, len(appro))
    assert frac > 0.90, f"FY2025 appropriated non-null only {frac:.1%} (<90%)"


def _run_account_balances_level(account_level: str, level_label: str, uri: str, btree_key: str,
                                *, so: dict, sweep, run_id, smoke: bool) -> dict:
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    stream = f"usa_account_balances:{level_label}"
    status, disp, notes = "failed", "none", ""
    tables: list = []
    per_fy: dict[int, int] = {}
    observed_headers: dict[int, list] = {}
    total_bytes = 0
    printed_header = False
    try:
        for (fy, quarter) in sweep:
            try:
                zb = _mint_and_download(account_level, fy, quarter)
            except ThrottleHalt:
                raise  # a breaker halt must propagate — never retry into a wall
            except (TimeoutError, RuntimeError) as exc:
                print(f"    RETRY once fy={fy} level={level_label}: {exc}", flush=True)
                zb = _mint_and_download(account_level, fy, quarter)  # one retry, else propagate
            total_bytes += len(zb)
            tbl, header = _parse_file_a(zb, fy, level_label)
            tables.append(tbl)
            per_fy[fy] = tbl.num_rows
            observed_headers[fy] = header
            if not printed_header:
                print(f"    {level_label} L44 header ({len(header)} cols): {header}", flush=True)
                printed_header = True
            print(f"    fy={fy} q={quarter} {level_label}: {tbl.num_rows:,} rows, "
                  f"{len(header)} cols", flush=True)
        combined = pa.concat_tables(tables, promote_options="default") if tables else None
        if combined is None:
            raise RuntimeError(f"no File-A tables landed for {level_label}")
        _write_lance(combined, uri, so)
        _build_indexes(uri, btree=[btree_key, "fiscal_year"], bitmap=[], so=so)
        notes = (f"per_fy_rows={json.dumps(per_fy)}; "
                 f"cols_by_fy={json.dumps({str(k): len(v) for k, v in observed_headers.items()})}")
        # federal-level gates run inline so a gate failure marks THIS ledger row 'failed'
        if level_label == "federal" and not smoke:
            _gate_file_a_federal(combined, per_fy, observed_headers)
        status, disp = "completed", "ok"
        return {"stream": stream, "rows": combined.num_rows, "status": status, "uri": uri,
                "per_fy": per_fy, "bytes": total_bytes}
    except ThrottleHalt:
        disp = "throttled"
        raise
    except Exception as exc:  # noqa: BLE001
        notes = f"{notes} ERROR: {exc}"[:8000]
        raise
    finally:
        _record_run(run_id, stream, USA_DOWNLOAD_ACCOUNTS, total_bytes or None,
                    sum(per_fy.values()), {level_label: per_fy}, t0,
                    dt.datetime.now(dt.timezone.utc), status, disp, notes)
        print(f"[{stream}] rows={sum(per_fy.values()):,} status={status} disp={disp}", flush=True)


def run_usa_account_balances(*, so: dict, smoke: bool, run_id, uris: dict) -> list[dict]:
    sweep = _resolve_account_sweep(smoke=smoke)
    print(f"[usa_account_balances] sweep={sweep}", flush=True)
    fed = _run_account_balances_level(
        "federal_account", "federal", uris["usaspending_federal_account_balances"],
        "federal_account_symbol", so=so, sweep=sweep, run_id=run_id, smoke=smoke)
    tre = _run_account_balances_level(
        "treasury_account", "treasury", uris["usaspending_treasury_account_balances"],
        "treasury_account_symbol", so=so, sweep=sweep, run_id=run_id, smoke=smoke)
    return [{k: r[k] for k in ("stream", "rows", "status", "uri")} for r in (fed, tre)]


# ═══════════════════════════════════════════════════════════════════════════════════
# Derived — approp_vs_obligated_fy (#10; reads omb_budget_authority, omb_outlays, File-A)
# ═══════════════════════════════════════════════════════════════════════════════════
COVERAGE_NOTE = (
    "USAspending budget_authority_appropriated_amount is a gross-discretionary basis on a "
    "TAS/DABS-submission footprint (submitting agencies only); compare it to omb_gross_ba_musd "
    "/ omb_discretionary_ba_musd, NEVER omb_net_ba_musd. OMB is $millions on a budget-account "
    "basis including offsetting receipts and off-budget accounts. usa_unobligated_balance_usd is "
    "the appropriated-but-unspent measure. The two systems do not tie out and are not meant to — "
    "different bases, different coverage; land both, label both, reconcile neither."
)


def run_derived(*, so: dict, uri: str, smoke: bool, run_id, uris: dict) -> dict:
    import duckdb
    import lance
    import pyarrow as pa

    t0 = dt.datetime.now(dt.timezone.utc)
    rows: list[dict] = []
    status, disp, notes, built = "failed", "none", "", []
    try:
        con = duckdb.connect(":memory:")

        def _load(name, u):
            try:
                con.register(name, lance.dataset(u, storage_options=so).to_table())
                return True
            except Exception as exc:  # noqa: BLE001
                print(f"  derived: {name} unavailable ({exc})", flush=True)
                return False

        have_ba = _load("omb_ba", uris["omb_budget_authority"])
        have_out = _load("omb_out", uris["omb_outlays"])
        have_fa = _load("fa", uris["usaspending_federal_account_balances"])
        if not (have_ba and have_fa):
            raise RuntimeError("derived requires omb_budget_authority and File-A federal balances")

        omb = con.execute("""
            SELECT year AS fiscal_year,
                   SUM(CASE WHEN row_kind='account' AND value_musd>0 THEN value_musd ELSE 0 END) AS omb_gross_ba_musd,
                   SUM(CASE WHEN value_musd<0 THEN value_musd ELSE 0 END) AS omb_offsetting_receipts_musd,
                   SUM(value_musd) AS omb_net_ba_musd,
                   SUM(CASE WHEN row_kind='account' AND value_musd>0 AND bea_category='Discretionary'
                            THEN value_musd ELSE 0 END) AS omb_discretionary_ba_musd,
                   BOOL_OR(is_estimate) AS is_estimate
            FROM omb_ba WHERE is_transition_quarter = FALSE
            GROUP BY year
        """).fetchall()
        omb_cols = [d[0] for d in con.description]
        omb_by_fy = {r[0]: dict(zip(omb_cols, r)) for r in omb}

        out_by_fy = {}
        if have_out:
            for r in con.execute("""
                SELECT year AS fiscal_year, SUM(value_musd) AS omb_outlays_musd
                FROM omb_out WHERE is_transition_quarter = FALSE GROUP BY year
            """).fetchall():
                out_by_fy[r[0]] = r[1]

        usa = con.execute("""
            SELECT fiscal_year,
                   SUM(budget_authority_appropriated_amount) AS usa_budget_authority_appropriated_usd,
                   SUM(total_budgetary_resources)            AS usa_total_budgetary_resources_usd,
                   SUM(obligations_incurred)                 AS usa_obligations_incurred_usd,
                   SUM(unobligated_balance)                  AS usa_unobligated_balance_usd,
                   SUM(gross_outlay_amount)                  AS usa_gross_outlay_usd
            FROM fa GROUP BY fiscal_year
        """).fetchall()
        usa_cols = [d[0] for d in con.description]
        usa_by_fy = {r[0]: dict(zip(usa_cols, r)) for r in usa}
        con.close()

        # gate: every USA fiscal_year must be present in OMB (fail on a one-sided year)
        for fy in usa_by_fy:
            if fy not in omb_by_fy:
                raise AssertionError(f"derived: USA FY{fy} absent from OMB budget authority")

        ingested_at = dt.datetime.now(dt.timezone.utc)
        all_fy = sorted(set(omb_by_fy) | set(usa_by_fy))
        for fy in all_fy:
            o = omb_by_fy.get(fy, {})
            u = usa_by_fy.get(fy, {})
            cov = "both" if (fy in omb_by_fy and fy in usa_by_fy) else \
                  ("omb_only" if fy in omb_by_fy else "usa_only")
            rows.append({
                "fiscal_year": int(fy),
                "omb_gross_ba_musd": _num(o.get("omb_gross_ba_musd")),
                "omb_offsetting_receipts_musd": _num(o.get("omb_offsetting_receipts_musd")),
                "omb_net_ba_musd": _num(o.get("omb_net_ba_musd")),
                "omb_discretionary_ba_musd": _num(o.get("omb_discretionary_ba_musd")),
                "omb_outlays_musd": _num(out_by_fy.get(fy)),
                "usa_budget_authority_appropriated_usd": _num(u.get("usa_budget_authority_appropriated_usd")),
                "usa_total_budgetary_resources_usd": _num(u.get("usa_total_budgetary_resources_usd")),
                "usa_obligations_incurred_usd": _num(u.get("usa_obligations_incurred_usd")),
                "usa_unobligated_balance_usd": _num(u.get("usa_unobligated_balance_usd")),
                "usa_gross_outlay_usd": _num(u.get("usa_gross_outlay_usd")),
                "is_estimate": bool(o.get("is_estimate")) if o else False,
                "coverage_note": COVERAGE_NOTE,
                "coverage": cov,
                "source": f"{SOURCE_TAG}:derived",
                "ingested_at": ingested_at,
            })
        schema = pa.schema([
            ("fiscal_year", pa.int32()),
            ("omb_gross_ba_musd", pa.float64()), ("omb_offsetting_receipts_musd", pa.float64()),
            ("omb_net_ba_musd", pa.float64()), ("omb_discretionary_ba_musd", pa.float64()),
            ("omb_outlays_musd", pa.float64()),
            ("usa_budget_authority_appropriated_usd", pa.float64()),
            ("usa_total_budgetary_resources_usd", pa.float64()),
            ("usa_obligations_incurred_usd", pa.float64()),
            ("usa_unobligated_balance_usd", pa.float64()),
            ("usa_gross_outlay_usd", pa.float64()),
            ("is_estimate", pa.bool_()), ("coverage_note", pa.string()),
            ("coverage", pa.string()), ("source", pa.string()),
            ("ingested_at", pa.timestamp("us", tz="UTC")),
        ])
        _write_lance(pa.Table.from_pylist(rows, schema=schema), uri, so)
        built = _build_indexes(uri, btree=["fiscal_year"], bitmap=[], so=so)
        notes = f"fy_rows={len(rows)}; usa_fys={sorted(usa_by_fy)}"
        status, disp = "completed", "ok"
    except ThrottleHalt:
        disp = "throttled"
        raise
    except Exception as exc:  # noqa: BLE001
        notes = f"{notes} ERROR: {exc}"[:8000]
        raise
    finally:
        _record_run(run_id, "derived", f"{SOURCE_TAG}:derived", None, len(rows),
                    {"approp_vs_obligated_fy": len(rows)}, t0,
                    dt.datetime.now(dt.timezone.utc), status, disp, notes)
        print(f"[derived] rows={len(rows):,} status={status} disp={disp}", flush=True)
    return {"stream": "derived", "rows": len(rows), "status": status, "uri": uri}


# ═══════════════════════════════════════════════════════════════════════════════════
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--stream", choices=[
        "omb_budauth", "omb_outlays", "usa_account_balances", "usa_agency",
        "usa_agency_periods", "usa_total_resources", "usa_account_roster", "usa_defc",
        "derived", "all"])
    ap.add_argument("--smoke", action="store_true",
                    help="throwaway URIs under smoke/; first 200 xlsx rows / 3 agencies / latest FY")
    ap.add_argument("--init-state", action="store_true",
                    help="apply ledger + catalog DDL and register sources, then exit")
    args = ap.parse_args()

    if args.init_state:
        apply_state_schema()
        return
    if not args.stream:
        ap.error("--stream is required (or use --init-state)")

    so = _storage_options()
    uris = dict(URIS)
    if args.smoke:
        uris = {k: v.replace("/active/", "/smoke/") for k, v in uris.items()}

    # register sources up front (idempotent) so the catalog reflects the run even if a
    # later stream fails; never fatal.
    try:
        dsn = _dsn()
        if dsn:
            import psycopg

            with psycopg.connect(dsn, autocommit=True) as conn, conn.cursor() as cur:
                cur.execute(LEDGER_DDL)
                cur.execute(CATALOG_DDL)
                _register_sources(cur)
    except Exception as exc:  # noqa: BLE001
        print(f"WARN: source registration failed: {exc}", flush=True)

    run_id = uuid.uuid4()
    print(f"=== federal_appropriations_ingest run_id={run_id} stream={args.stream} "
          f"smoke={args.smoke} ===", flush=True)
    omb_urls = None

    def _omb(stream, key):
        nonlocal omb_urls
        if omb_urls is None:
            omb_urls = _resolve_omb_urls()
        url = omb_urls["budauth" if stream == "omb_budauth" else "outlays"]
        return _run_omb(stream, uris[key], url, so=so, smoke=args.smoke, run_id=run_id)

    # (stream, thunk). ThrottleHalt propagates to the breaker handler; any OTHER exception is
    # isolated to its stream (its own finally already recorded status='failed') so later
    # streams still run and record their ledger row.
    dispatch = [
        ("omb_budauth", lambda: _omb("omb_budauth", "omb_budget_authority")),
        ("omb_outlays", lambda: _omb("omb_outlays", "omb_outlays")),
        ("usa_agency", lambda: run_usa_agency(
            so=so, uri=uris["usaspending_agency_fy"], smoke=args.smoke, run_id=run_id)),
        ("usa_agency_periods", lambda: run_usa_agency_periods(
            so=so, uri=uris["usaspending_agency_period_obligations"], smoke=args.smoke, run_id=run_id)),
        ("usa_account_balances", lambda: run_usa_account_balances(
            so=so, smoke=args.smoke, run_id=run_id, uris=uris)),
        ("usa_total_resources", lambda: run_usa_total_resources(
            so=so, uri=uris["usaspending_total_budgetary_resources"], smoke=args.smoke, run_id=run_id)),
        ("usa_account_roster", lambda: run_usa_account_roster(
            so=so, uri=uris["usaspending_federal_account_roster"], smoke=args.smoke, run_id=run_id)),
        ("usa_defc", lambda: run_usa_defc(
            so=so, uri=uris["usaspending_def_codes"], smoke=args.smoke, run_id=run_id)),
        ("derived", lambda: run_derived(
            so=so, uri=uris["approp_vs_obligated_fy"], smoke=args.smoke, run_id=run_id, uris=uris)),
    ]

    results: list[dict] = []
    s = args.stream
    try:
        for name, thunk in dispatch:
            if s not in (name, "all"):
                continue
            try:
                r = thunk()
                results.extend(r if isinstance(r, list) else [r])
            except ThrottleHalt:
                raise
            except Exception as exc:  # noqa: BLE001 — isolate the stream; keep the batch going
                import traceback

                print(f"[{name}] STREAM FAILED: {exc}", flush=True)
                traceback.print_exc()
                results.append({"stream": name, "rows": 0, "status": "failed",
                                "uri": uris.get(name, "")})
    except ThrottleHalt as exc:
        for g in _GOVS.values():
            g.flush_checkpoint()
        print(f"\n!!! CIRCUIT-BREAKER HALT on {exc.host}: {exc}. Run stopped; "
              f"disposition=throttled. Resume from checkpoint after cool-down.", flush=True)
        raise SystemExit(2)
    finally:
        for g in _GOVS.values():
            g.flush_checkpoint()

    print("\n=== BATCH RESULT ===", flush=True)
    for r in results:
        print(f"  {r['stream']:28s} rows={r['rows']:>8,} status={r['status']}", flush=True)
    if any(r["status"] != "completed" for r in results):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
