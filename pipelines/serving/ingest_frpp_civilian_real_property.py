"""Reference ingest — `frpp_civilian_real_property` (Federal Real Property Profile, FY24 public).

SoR  s3://data-sink/active/frpp_civilian_real_property/  (Lance v2.1; reference, snapshot-overwrite).

WHAT THIS IS
The wide net around `gsa_buildings_lance`: one row per federal civilian real-property asset in
the FY2024 FRPP public dataset — ~308k assets across every reporting agency (not just GSA),
each carrying rooftop `Latitude`/`Longitude` (the public civilian release is NOT
coordinate-redacted), reporting/using agency, real-property type/use, square footage,
replacement value, condition index, and admin geo (state/county/city/zip/congressional
district). Point geometry (POINT WKT) is emitted wherever lat/lon parse, so the same
haversine-proximity path used for `gsa_buildings_lance` works here over the full federal footprint.

SOURCE (GSA FRPP Public Dataset, FY2024, civilian agencies — single XLSX, ~139 MB)
  https://inventory.data.gov/dataset/5752ee7f-9e8b-467a-aa5a-274b4bd1bc29/resource/
    1a94a302-b3c9-433d-ac97-43db73fa6d04/download/frpp_public_dataset_fy24_07022025.xlsx
  Sheet `sheet1`, 117 verbatim columns, header row 1.

NORMALIZATION
  * Column names: snake_case slug of each verbatim header (117 cols), de-duplicated.
  * `latitude`/`longitude` are typed DOUBLE in place; every other source column is carried
    verbatim as a source-faithful string (empty/whitespace → NULL). Money ($-formatted) and
    comma-grouped measures stay as source strings — downstream recipes cast at read time.
  * Geometry: POINT(lon lat) WKT wherever lat/lon parse; consumers use ST_GeomFromText.

GRAIN: 1 row / (real_property_unique_identifier) as reported for FY2024 (distinctness reported
at build; RPUID BTREE is non-unique-safe). Idempotent snapshot-overwrite.
  BTREE(real_property_unique_identifier, installation_id) · BITMAP(reporting_agency_code,
  using_agency_code, state_code, real_property_type_code, real_property_use_code,
  legal_interest_code, asset_status_code, us_foreign).

    doppler run --project core-x --config prd -- python pipelines/serving/ingest_frpp_civilian_real_property.py
    doppler run --project core-x --config prd -- python pipelines/serving/ingest_frpp_civilian_real_property.py --verify
"""
from __future__ import annotations

import datetime as dt
import os
import re
import sys
import urllib.request

A = "s3://data-sink/active"
SRC_URL = ("https://inventory.data.gov/dataset/5752ee7f-9e8b-467a-aa5a-274b4bd1bc29/resource/"
           "1a94a302-b3c9-433d-ac97-43db73fa6d04/download/frpp_public_dataset_fy24_07022025.xlsx")
SERVING_URI = os.environ.get("FRPP_CIVILIAN_LANCE_URI", f"{A}/frpp_civilian_real_property/")
DATA_STORAGE_VERSION = "2.1"
SOURCE_VERSION = "frpp_public_dataset_fy24"
SHEET = "sheet1"
EXPECTED_COLS = 117
BATCH_ROWS = 20000
TMP_XLSX = "/tmp/frpp_fy24.xlsx"

FLOAT_COLS = {"latitude", "longitude"}  # typed in place; all other source cols stay string
BTREE_COLS = ["real_property_unique_identifier", "installation_id"]
BITMAP_COLS = ["reporting_agency_code", "using_agency_code", "state_code",
               "real_property_type_code", "real_property_use_code", "legal_interest_code",
               "asset_status_code", "us_foreign"]


def _r2_storage_options() -> dict[str, str]:
    endpoint = os.environ.get("R2_ENDPOINT")
    account_id = os.environ.get("R2_ACCOUNT_ID")
    if not endpoint and account_id:
        endpoint = f"https://{account_id}.r2.cloudflarestorage.com"
    if not endpoint:
        raise RuntimeError("Set R2_ENDPOINT or R2_ACCOUNT_ID.")
    return {"aws_access_key_id": os.environ["R2_ACCESS_KEY_ID"],
            "aws_secret_access_key": os.environ["R2_SECRET_ACCESS_KEY"],
            "endpoint": endpoint, "region": "auto"}


def _slug(header: str) -> str:
    s = re.sub(r"[^0-9a-zA-Z]+", "_", (header or "").strip().lower())
    return s.strip("_") or "col"


def _dedup(slugs: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for s in slugs:
        if s in seen:
            seen[s] += 1
            out.append(f"{s}_{seen[s]}")
        else:
            seen[s] = 0
            out.append(s)
    return out


def _clean_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_float(v) -> float | None:
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def _download(path: str = TMP_XLSX) -> str:
    req = urllib.request.Request(SRC_URL, headers={"User-Agent": "core-x-ingest/1.0"})
    with urllib.request.urlopen(req, timeout=600) as r, open(path, "wb") as f:
        while True:
            chunk = r.read(1 << 20)
            if not chunk:
                break
            f.write(chunk)
    return path


def _open_sheet(xlsx_path: str):
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    return wb, ws


def build(local_xlsx: str | None = None) -> dict:
    import lance
    import pyarrow as pa

    xlsx = local_xlsx or _download()

    # ── header pass: normalize + validate the column contract ──
    wb, ws = _open_sheet(xlsx)
    rows_it = ws.iter_rows(values_only=True)
    header = next(rows_it)
    if len(header) != EXPECTED_COLS:
        raise RuntimeError(f"schema drift: {len(header)} cols != expected {EXPECTED_COLS}")
    slugs = _dedup([_slug(h) for h in header])
    for required in ("real_property_unique_identifier", "latitude", "longitude"):
        if required not in slugs:
            raise RuntimeError(f"missing expected column after slug: {required}")

    out_cols = list(slugs) + ["geometry_wkt", "geometry_type", "source_version", "materialized_at"]
    schema = pa.schema(
        [pa.field(s, pa.float64() if s in FLOAT_COLS else pa.string()) for s in slugs]
        + [pa.field("geometry_wkt", pa.string()),
           pa.field("geometry_type", pa.string()),
           pa.field("source_version", pa.string(), nullable=False),
           pa.field("materialized_at", pa.timestamp("us", tz="UTC"), nullable=False)]
    )

    now = dt.datetime.now(dt.timezone.utc)
    stats = {"rows": 0, "point_geoms": 0}
    rpuid_idx = slugs.index("real_property_unique_identifier")
    seen_rpuid: set[str] = set()

    def _batches():
        buf: dict[str, list] = {c: [] for c in out_cols}
        n = 0
        for row in rows_it:
            rec: dict = {}
            for s, val in zip(slugs, row):
                rec[s] = _to_float(val) if s in FLOAT_COLS else _clean_str(val)
            lon, lat = rec.get("longitude"), rec.get("latitude")
            if lon is not None and lat is not None:
                rec["geometry_wkt"] = f"POINT ({lon} {lat})"
                rec["geometry_type"] = "Point"
                stats["point_geoms"] += 1
            else:
                rec["geometry_wkt"] = None
                rec["geometry_type"] = None
            rec["source_version"] = SOURCE_VERSION
            rec["materialized_at"] = now
            rid = row[rpuid_idx]
            if rid is not None:
                seen_rpuid.add(str(rid).strip())
            for c in out_cols:
                buf[c].append(rec.get(c))
            n += 1
            stats["rows"] += 1
            if n >= BATCH_ROWS:
                yield pa.record_batch({c: buf[c] for c in out_cols}, schema=schema)
                buf = {c: [] for c in out_cols}
                n = 0
        if n:
            yield pa.record_batch({c: buf[c] for c in out_cols}, schema=schema)

    reader = pa.RecordBatchReader.from_batches(schema, _batches())
    so = _r2_storage_options()
    lance.write_dataset(reader, SERVING_URI, mode="overwrite",
                        data_storage_version=DATA_STORAGE_VERSION, storage_options=so)
    wb.close()

    rows = stats["rows"]
    print(f"parsed {rows:,} FRPP assets "
          f"(distinct rpuid: {len(seen_rpuid):,} · point geometry: {stats['point_geoms']:,})")
    if rows == 0:
        raise RuntimeError("no rows written")

    ds = lance.dataset(SERVING_URI, storage_options=so)
    present = set(ds.schema.names)
    for c in BTREE_COLS:
        if c in present:
            try:
                ds.create_scalar_index(c, index_type="BTREE"); print(f"  BTREE ✓ {c}")
            except Exception as exc:  # noqa: BLE001
                print(f"  WARN BTREE {c}: {exc}")
    for c in BITMAP_COLS:
        if c in present:
            try:
                ds.create_scalar_index(c, index_type="BITMAP"); print(f"  BITMAP ✓ {c}")
            except Exception as exc:  # noqa: BLE001
                print(f"  WARN BITMAP {c}: {exc}")
    back = ds.count_rows()
    assert back == rows, f"write-integrity gate: {back} != {rows}"
    print(f"WROTE {SERVING_URI} rows={back} cols={len(ds.schema)}")
    return {"uri": SERVING_URI, "rows": back,
            "distinct_rpuid": len(seen_rpuid), "point_geoms": stats["point_geoms"]}


def verify() -> None:
    import duckdb
    import lance

    so = _r2_storage_options()
    ds = lance.dataset(SERVING_URI, storage_options=so)
    print(f"{SERVING_URI}  rows={ds.count_rows():,}  cols={len(ds.schema)}")
    print("indices:", sorted(
        (i.get("name") if isinstance(i, dict) else getattr(i, "name", str(i))) for i in ds.list_indices()))
    con = duckdb.connect(":memory:"); con.register("f", ds)
    print("\n=== coverage ===")
    print(con.execute("""SELECT
        count(*) total,
        count(DISTINCT real_property_unique_identifier) distinct_rpuid,
        count(*) FILTER (WHERE geometry_wkt IS NOT NULL) with_point,
        count(*) FILTER (WHERE us_foreign = 'UNITED STATES') us_rows
    FROM f""").df().to_string(index=False))
    print("\n=== top 12 reporting agencies ===")
    print(con.execute("SELECT reporting_agency, count(*) n FROM f "
                      "GROUP BY 1 ORDER BY 2 DESC LIMIT 12").df().to_string(index=False))
    print("\n=== real property type mix (top 10) ===")
    print(con.execute("SELECT real_property_type, count(*) n FROM f "
                      "GROUP BY 1 ORDER BY 2 DESC LIMIT 10").df().to_string(index=False))
    print("\n=== legal interest (owned vs leased) ===")
    print(con.execute("SELECT legal_interest_indicator, count(*) n FROM f "
                      "GROUP BY 1 ORDER BY 2 DESC").df().to_string(index=False))
    try:
        con.execute("INSTALL spatial; LOAD spatial;")
        r = con.execute("SELECT real_property_unique_identifier, installation_name, "
                        "ST_GeometryType(ST_GeomFromText(geometry_wkt)) gtype "
                        "FROM f WHERE geometry_wkt IS NOT NULL LIMIT 1").df()
        print("\n=== spatial round-trip ===")
        print(r.to_string(index=False))
    except Exception as exc:  # noqa: BLE001
        print(f"WARN spatial round-trip: {exc}")


if __name__ == "__main__":
    if "--verify" in sys.argv:
        verify()
    else:
        lz = next((a for a in sys.argv[1:] if not a.startswith("-")), None)
        build(local_xlsx=lz)
